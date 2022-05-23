import pyodbc 
import os
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
import pandas as pd
import chardet
from xlsxwriter import Workbook
import glob

from openpyxl import formatting, styles, Workbook as openpyxl_workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill, colors
from openpyxl.styles.colors import Color, ColorDescriptor
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.styles.fills import Fill
from openpyxl.formatting.rule import CellIsRule, Rule
from openpyxl.styles.borders import Border, Side
import time


cwd = os.path.dirname(os.path.realpath(__file__))
os.chdir(cwd)

#'Server=localhost\SQLEXPRESS;'
conn = pyodbc.connect('Driver={ODBC Driver 17 for SQL Server};'
                      'Server=localhost;'
                      'Database=DashboardData;'
                      'Trusted_Connection=yes;')

cursor = conn.cursor()
def change_format(date):
    #print('date', date)
    date = date.strip()
    
    if date:
        
        try:
            return datetime.strptime(date, "%d-%b-%y").strftime('%Y-%m-%d')
        except:
            try: 
                return datetime.strptime(date, "%d-%b-%Y").strftime('%Y-%m-%d')
            except: 
                try:
                    return datetime.strptime(date, "%m/%d/%Y").strftime('%Y-%m-%d')
                except: 
                    
                    return datetime.strptime('', "%m/%d/%Y").strftime('%Y-%m-%d')
    else:
        return date

def change_display_format(date):
    date = date.strip()
    if date:
        try:
            return datetime.strptime(date, "%Y-%m-%d").strftime('%d-%b-%y')
        except:
            return date
    
def truncate_all():
    cursor.execute("Delete from [Case];\
    Delete from [Beneficiary];\
    truncate table [BeneficiaryPriorityDate];\
    truncate table [BeneficiaryEmployment];\
    truncate table [Petitioner];\
    truncate table [Organization];")
    cursor.commit()
    
    

def start():
    truncate_all()
    current_time = datetime.now() 
    month = str(current_time.month).rjust(2, '0')
    day = str(current_time.day).rjust(2, '0')
    todate = month+''+day+''+str(current_time.year)
    from_name = ''
    for name in glob.glob('Source Data Dashboard/*Case*'):
        ##print(os.path.basename(name))
        filename_e = os.path.basename(name)
        filename = os.path.splitext(filename_e)[0]
        extension = os.path.splitext(filename_e)[1]
        ##print(extension)
        if extension == '.csv' and  todate in filename:
            from_name = (filename.split('Data_'))[1].split('_'+str(todate))[0].strip()
            case_file_name = 'Practice Management Dashboard_Case Data_'+str(from_name)+'_'+todate+'.csv'
            ##print(benificiary_file_name)
            if os.path.exists('Source Data Dashboard/'+case_file_name):
                print('Processing - '+case_file_name)
                process_case_file('Source Data Dashboard/'+case_file_name, from_name)

    if from_name:
        print('Generating Report - '+from_name)
        generate_case_report(todate)
        
        
def process_case_file(file_path, from_name):
    with open(file_path,'rb') as f:
        rawdata = b''.join([f.readline() for _ in range(20)])
    enc= chardet.detect(rawdata)['encoding'] #UTF-16

    df = pd.read_csv(file_path, encoding=enc,delimiter='\t')
    list_h = df.columns.tolist()
    total_rows = len(df)
    for index, row in df.iterrows():
        organization_xref = ''
        if 'Organization Xref' in list_h:
            organization_xref = row['Organization Xref'].strip()
        
        organization_name = ''
        if "Organization Name" in list_h:
            organization_name = (row['Organization Name'].replace("'", "")).strip()

        organization_id = ''
        if organization_xref  and organization_name :
            results = cursor.execute("SELECT * FROM dbo.Organization where OrganizationXref='{}' and OrganizationName = '{}'".format(organization_xref, organization_name)).fetchall()
            length = len(results)
            if length <= 0:
                cursor.execute("INSERT INTO dbo.Organization(OrganizationXref, OrganizationName) VALUES ('{}', '{}')".format(organization_xref, organization_name))
                cursor.execute("SELECT @@IDENTITY AS ID;")
                organization_id = cursor.fetchone()[0]
                cursor.commit()
                
            else:
                organization_id = results[0].OrganizationId
        
        
        petitioner_xref = ''
        if "Petitioner Xref" in list_h:
            petitioner_xref = row['Petitioner Xref'].strip()
        
        petitioner_name = ''
        if "Petitioner Name" in list_h:
            petitioner_name = (row['Petitioner Name'].replace("'", "")).strip()

        petitioner_id = ''
        is_primary_beneficiary = 1
        if petitioner_xref and petitioner_name:
            if petitioner_name == 'Individual Client' :
                if row['Primary Beneficiary Xref'].strip():
                    results = cursor.execute("SELECT PetitionerId FROM dbo.Beneficiary where BeneficiaryXref='{}'".format(row['Primary Beneficiary Xref'].strip())).fetchall()
                    length = len(results)
                    if length > 0:
                        petitioner_id = results[0][0]
                        is_primary_beneficiary = 0
                   
                    

            else:
                results = cursor.execute("SELECT * FROM dbo.Petitioner where PetitionerXref='{}' and PetitionerName = '{}' and OrganizationId={}".format(petitioner_xref, petitioner_name, organization_id)).fetchall()
                length = len(results)
                if length <= 0:
                    cursor.execute("INSERT INTO dbo.Petitioner(PetitionerXref, PetitionerName, OrganizationId) VALUES ('{}', '{}', '{}')".format(petitioner_xref, petitioner_name, organization_id))
                    cursor.execute("SELECT @@IDENTITY AS ID;")
                    petitioner_id = cursor.fetchone()[0]
                    cursor.commit()
                else:
                    petitioner_id = results[0].PetitionerId
        
        #if petitioner_id :
        if True:
            beneficiary_xref = ''
            if "Beneficiary Xref" in list_h and not pd.isna(row["Beneficiary Xref"]):
                beneficiary_xref = row["Beneficiary Xref"]
            
            beneficiary_type = ''
            if "Beneficiary Type" in list_h and not pd.isna(row["Beneficiary Type"]):
                beneficiary_type = row["Beneficiary Type"]
            
            beneficiary_record_creation_date = ''
            if "Beneficiary Record Creation Date" in list_h and row["Beneficiary Record Creation Date"].strip() and not pd.isna(row["Beneficiary Record Creation Date"]):
                beneficiary_record_creation_date = change_format(row["Beneficiary Record Creation Date"])
            
            beneficiary_record_inactivation_date = ''
            if "Beneficiary Record Inactivation Date" in list_h and row["Beneficiary Record Inactivation Date"].strip() and not pd.isna(row["Beneficiary Record Inactivation Date"]):
                beneficiary_record_inactivation_date = change_format(row["Beneficiary Record Inactivation Date"])

            beneficiary_record_status = 0
            if "Beneficiary Record Status" in list_h and not pd.isna(row["Beneficiary Record Status"]):
                beneficiary_record_status = row["Beneficiary Record Status"]
                if beneficiary_record_status == 'Active':
                    beneficiary_record_status = 1
                else:
                    beneficiary_record_status = 0

            beneficiary_last_name = ''
            if "Beneficiary Last Name" in list_h and not pd.isna(row["Beneficiary Last Name"]):
                beneficiary_last_name = row["Beneficiary Last Name"].replace("'", "")

            beneficiary_first_name = ''
            if "Beneficiary First Name" in list_h  and not pd.isna(row["Beneficiary First Name"]):
                beneficiary_first_name = row["Beneficiary First Name"].replace("'", "")

            beneficiary_middle_name = ''
            if "Beneficiary Middle Name" in list_h and not pd.isna(row["Beneficiary Middle Name"]):
                beneficiary_middle_name = row["Beneficiary Middle Name"].replace("'", "")

            primary_beneficiary_id = ''
            if "Primary Beneficiary Xref" in list_h and not pd.isna(row["Primary Beneficiary Xref"]):
                primary_beneficiary_id = row["Primary Beneficiary Xref"]

            if primary_beneficiary_id == beneficiary_xref:
                is_primary_beneficiary = 1
            else:
                is_primary_beneficiary = 0

            primary_beneficiary_last_name = ''
            if "Primary Beneficiary Last Name" in list_h and not pd.isna(row["Primary Beneficiary Last Name"]):
                primary_beneficiary_last_name = row["Primary Beneficiary Last Name"].replace("'", "")
            
            primary_beneficiary_first_name = ''
            if "Primary Beneficiary First Name" in list_h and not pd.isna(row["Primary Beneficiary First Name"]):
                primary_beneficiary_first_name = row["Primary Beneficiary First Name"].replace("'", "")
            
            relation = ''
            if "Relation" in list_h and not pd.isna(row["Relation"]):
                relation = row["Relation"]

            immigration_status = ''
            if "Immigration Status" in list_h and not pd.isna(row["Immigration Status"]):
                immigration_status = row["Immigration Status"]

            immigration_status_expiration_status = ''
            if "Immigration Status Expiration Date" in list_h and row["Immigration Status Expiration Date"].strip() and not pd.isna(row["Immigration Status Expiration Date"]):
                if row["Immigration Status Expiration Date"].strip() == 'D/S':
                    immigration_status_expiration_status = 'D/S'
                else:
                    if 'D/S' in row["Immigration Status Expiration Date"].strip():
                        split1 = (row["Immigration Status Expiration Date"].strip()).split('(D/S)')
                        immigration_status_expiration_status = change_format(split1[0])
                        immigration_status_expiration_status = str(immigration_status_expiration_status)+' (D/S)'
                    else:
                        immigration_status_expiration_status = change_format(row["Immigration Status Expiration Date"])

            i797_approved_date = ''
            if "I-797 Approved Date" in list_h and row["I-797 Approved Date"].strip() and not pd.isna(row["I-797 Approved Date"]):
                i797_approved_date = change_format(row["I-797 Approved Date"])

            i797_status = ''
            if "I-797 Status" in list_h and not pd.isna(row["I-797 Status"]):
                i797_status = row["I-797 Status"]
            
            i797_expiration_date = ''
            if "I-797 Expiration Date" in list_h and row["I-797 Expiration Date"].strip() and not pd.isna(row["I-797 Expiration Date"]):
                i797_expiration_date = change_format(row["I-797 Expiration Date"])

            final_niv_maxout_date = ''
            if "Final NIV (Maxout) Date" in list_h and row["Final NIV (Maxout) Date"].strip() and not pd.isna(row["Final NIV (Maxout) Date"]):
                final_niv_maxout_date = change_format(row["Final NIV (Maxout) Date"])

            maxout_note = ''
            if "Maxout Date Applicability and Note" in list_h and not pd.isna(row["Maxout Date Applicability and Note"]):
                maxout_note = row["Maxout Date Applicability and Note"]

            beneficiary_id = ''
            if beneficiary_xref:
                results = cursor.execute("SELECT * FROM dbo.Beneficiary where BeneficiaryXref='{}' and from_name='{}'".format(beneficiary_xref, from_name)).fetchall()
                length = len(results)
                if length <= 0:
                    
                    cursor.execute("INSERT INTO dbo.Beneficiary(PetitionerId, BeneficiaryXref, BeneficiaryType, SourceCreatedDate, IsActive, InactiveDate, LastName, FirstName, MiddleName, PrimaryBeneficiaryXref, PrimaryBeneficiaryLastName, PrimaryBeneficiaryFirstName, RelationType, ImmigrationStatus, ImmigrationStatusExpirationDate, MostRecentI797IssueApprovalDate, MostRecentI797Status, I797ExpirationDate, FinalNivDate, MaxOutDateNote, from_name, is_primary_beneficiary  ) VALUES ('{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}','{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}')".format(petitioner_id, beneficiary_xref, beneficiary_type, beneficiary_record_creation_date, beneficiary_record_status, beneficiary_record_inactivation_date, beneficiary_last_name, beneficiary_first_name, beneficiary_middle_name, primary_beneficiary_id, primary_beneficiary_last_name, primary_beneficiary_first_name, relation, immigration_status, immigration_status_expiration_status, i797_approved_date, i797_status, i797_expiration_date, final_niv_maxout_date, maxout_note, from_name, is_primary_beneficiary))
                    cursor.execute("SELECT @@IDENTITY AS ID;")
                    beneficiary_id = cursor.fetchone()[0]
                    cursor.commit()
                else:
                    beneficiary_id = results[0].BeneficiaryId
                    cursor.execute("UPDATE  dbo.Beneficiary SET PetitionerId='{}', BeneficiaryXref='{}', BeneficiaryType='{}', SourceCreatedDate='{}', IsActive='{}', InactiveDate='{}', LastName='{}', FirstName='{}', MiddleName='{}', PrimaryBeneficiaryXref='{}', PrimaryBeneficiaryLastName='{}', PrimaryBeneficiaryFirstName='{}', RelationType='{}', ImmigrationStatus='{}', ImmigrationStatusExpirationDate='{}', MostRecentI797IssueApprovalDate='{}', MostRecentI797Status='{}', I797ExpirationDate='{}', FinalNivDate='{}', MaxOutDateNote='{}', from_name='{}', is_primary_beneficiary='{}' WHERE BeneficiaryId='{}'  ".format(petitioner_id, beneficiary_xref, beneficiary_type, beneficiary_record_creation_date, beneficiary_record_status, beneficiary_record_inactivation_date, beneficiary_last_name, beneficiary_first_name, beneficiary_middle_name, primary_beneficiary_id, primary_beneficiary_last_name, primary_beneficiary_first_name, relation, immigration_status, immigration_status_expiration_status, i797_approved_date, i797_status, i797_expiration_date, final_niv_maxout_date, maxout_note, from_name, is_primary_beneficiary, beneficiary_id))
                    cursor.commit()

            if beneficiary_id:
                case_xref = ''
                if "Case Xref" in list_h and not pd.isna(row["Case Xref"]):
                    case_xref = row["Case Xref"]
                
                case_creation_date = ''
                if "Case Created Date" in list_h and row["Case Created Date"].strip() and not pd.isna(row["Case Created Date"]):
                    case_creation_date = change_format(row["Case Created Date"])

                case_petition_name = ''
                if "Case Petition Name" in list_h and not pd.isna(row["Case Petition Name"]):
                    case_petition_name = row["Case Petition Name"].replace("'", "")

                case_type = ''
                if "Case Type" in list_h and not pd.isna(row["Case Type"]):
                    case_type = row["Case Type"].replace("'", "")

                case_description = ''
                if "Case Description" in list_h and not pd.isna(row["Case Description"]):
                    case_description = row["Case Description"].replace("'", "")
                
                case_filed_date = ''
                if "Case Filed Date" in list_h and row["Case Filed Date"].strip() and not pd.isna(row["Case Filed Date"]):
                    case_filed_date = change_format(row["Case Filed Date"])
                
                
                case_receipt_number = ''
                if "Case Receipt Number" in list_h and not pd.isna(row["Case Receipt Number"]):
                    case_receipt_number = row["Case Receipt Number"]

                case_receipt_status = ''
                if "Case Receipt Status" in list_h and not pd.isna(row["Case Receipt Status"]):
                    case_receipt_status = row["Case Receipt Status"]

                rfe_audit_received_date = ''
                if "RFE/Audit Received Date" in list_h and row["RFE/Audit Received Date"].strip() and not pd.isna(row["RFE/Audit Received Date"]):
                    rfe_audit_received_date = change_format(row["RFE/Audit Received Date"])
                
                rfe_audit_due_date = ''
                if "RFE/Audit Response Due Date" in list_h and row["RFE/Audit Response Due Date"].strip() and not pd.isna(row["RFE/Audit Response Due Date"]):
                    rfe_audit_due_date = change_format(row["RFE/Audit Response Due Date"])
                
                rfe_audit_submitted_date = ''
                if "RFE/Audit Response Submitted Date" in list_h and row["RFE/Audit Response Submitted Date"].strip() and not pd.isna(row["RFE/Audit Response Submitted Date"]):
                    rfe_audit_submitted_date = change_format(row["RFE/Audit Response Submitted Date"])

                primary_case_status = ''
                if "Primary Case Status" in list_h and not pd.isna(row["Primary Case Status"]):
                    primary_case_status = row["Primary Case Status"]

                secondary_case_status = ''
                if "Secondary Case Status" in list_h and not pd.isna(row["Secondary Case Status"]):
                    secondary_case_status = row["Secondary Case Status"].replace("'", "")
                
                case_comments = ''
                if "Case Comments" in list_h and not pd.isna(row["Case Comments"]):
                    case_comments = row["Case Comments"].replace("'", "")

                case_last_step_completed = ''
                if "Case Last Step Completed" in list_h and not pd.isna(row["Case Last Step Completed"]):
                    case_last_step_completed = row["Case Last Step Completed"].replace("'", "")
                    case_last_step_completed = case_last_step_completed.replace("'", "`")

                case_last_step_completed_date = ''
                if "Case Last Step Completed Date" in list_h and row["Case Last Step Completed Date"].strip() and not pd.isna(row["Case Last Step Completed Date"]):
                    case_last_step_completed_date = change_format(row["Case Last Step Completed Date"])

                case_next_step_to_be_completed = ''
                if "Case Next Step To Be Completed" in list_h and not pd.isna(row["Case Next Step To Be Completed"]):
                    case_next_step_to_be_completed = row["Case Next Step To Be Completed"].replace("'", "")
                
                case_next_step_to_be_completed_date = ''
                if "Case Next Step To Be Completed Date" in list_h and row["Case Next Step To Be Completed Date"].strip() and not pd.isna(row["Case Next Step To Be Completed Date"]):
                    case_next_step_to_be_completed_date = change_format(row["Case Next Step To Be Completed Date"])
                
                case_priority_date = ''
                if "Case Priority Date" in list_h and row["Case Priority Date"].strip() and not pd.isna(row["Case Priority Date"]):
                    case_priority_date = change_format(row["Case Priority Date"])

                case_priority_category = ''
                if "Case Priority Category" in list_h and not pd.isna(row["Case Priority Category"]):
                    case_priority_category = row["Case Priority Category"]

                case_priority_country = ''
                if "Case Priority Country" in list_h and not pd.isna(row["Case Priority Country"]):
                    case_priority_country = row["Case Priority Country"]

                case_approved_date = '' 
                if "Case Approved Date" in list_h and row["Case Approved Date"].strip() and not pd.isna(row["Case Approved Date"]):
                    case_approved_date = change_format(row["Case Approved Date"])
                
                case_valid_from = ''
                if "Case Valid From" in list_h and row["Case Valid From"].strip() and not pd.isna(row["Case Valid From"]):
                    case_valid_from = change_format(row["Case Valid From"])
                
                case_valid_to = ''
                if "Case Valid To" in list_h and row["Case Valid To"].strip() and not pd.isna(row["Case Valid To"]):
                    case_valid_to = change_format(row["Case Valid To"])
                
                case_closed_date = ''
                if "Case Closed Date" in list_h and row["Case Closed Date"].strip() and not pd.isna(row["Case Closed Date"]):
                    case_closed_date = change_format(row["Case Closed Date"])
                
                case_denied_date = ''
                if "Case Denied Date" in list_h and row["Case Denied Date"].strip() and not pd.isna(row["Case Denied Date"]):
                    case_denied_date = change_format(row["Case Denied Date"])
                
                case_withdrawn_date = ''
                if "Case Withdrawn Date" in list_h and row["Case Withdrawn Date"].strip() and not pd.isna(row["Case Withdrawn Date"]):
                    case_withdrawn_date = change_format(row["Case Withdrawn Date"])
                
                case_primary_attorney = ''
                if "Case Primary Attorney" in list_h and not pd.isna(row["Case Primary Attorney"]):
                    case_primary_attorney = row["Case Primary Attorney"].replace("'", "")
                
                case_reviewing_attorney = ''
                if "Case Reviewing Attorney" in list_h and not pd.isna(row["Case Reviewing Attorney"]):
                    case_reviewing_attorney = row["Case Reviewing Attorney"].replace("'", "")
                
                case_primary_case_manager = ''
                if "Case Primary Case Manager" in list_h and not pd.isna(row["Case Primary Case Manager"]):
                    case_primary_case_manager = row["Case Primary Case Manager"].replace("'", "")
                
                petition_xref = ''
                if "Petition Xref" in list_h and not pd.isna(row["Petition Xref"]):
                    petition_xref = row["Petition Xref"]
                
                case_id = ''
                if case_xref:
                    
                    results = cursor.execute("SELECT * FROM [dbo].[Case] where BeneficiaryId='{}' and CaseXref='{}' and from_name='{}'".format(beneficiary_id, case_xref, from_name)).fetchall()
                    length = len(results)
                    if length <= 0:
                        cursor.execute("INSERT INTO [dbo].[Case](CaseXref, BeneficiaryId, SourceCreatedDate, CasePetitionName, CaseType, CaseDescription, CaseFiledDate, ReceiptNumber, ReceiptStatus, RFEAuditReceivedDate,RFEAuditDueDate, RFEAuditSubmittedDate, PrimaryCaseStatus, SecondaryCaseStatus, CaseComments, LastStepCompleted, LastStepCompletedDate, NextStepAction, NextStepActionDueDate, PriorityDate, PriorityCategory, PriorityCountry, CaseApprovedDate, CaseValidFromDate, CaseExpirationDate, CaseClosedDate, CaseDeniedDate, CaseWithdrawnDate, CasePrimaryAttorney, CaseReviewingAttorney, CasePrimaryCaseManager, PetitionXref, from_name) VALUES ('{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}','{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}')".format(case_xref, beneficiary_id, case_creation_date, case_petition_name, case_type, case_description, case_filed_date, case_receipt_number, case_receipt_status, rfe_audit_received_date, rfe_audit_due_date, rfe_audit_submitted_date, primary_case_status, secondary_case_status, case_comments, case_last_step_completed, case_last_step_completed_date, case_next_step_to_be_completed, case_next_step_to_be_completed_date, case_priority_date, case_priority_category, case_priority_country, case_approved_date, case_valid_from, case_valid_to, case_closed_date, case_denied_date, case_withdrawn_date, case_primary_attorney, case_reviewing_attorney, case_primary_case_manager, petition_xref, from_name))
                        cursor.execute("SELECT @@IDENTITY AS ID;")
                        case_id = cursor.fetchone()[0]
                        cursor.commit()
                    else:
                        case_id = results[0].CaseId
                        cursor.execute("UPDATE [dbo].[Case] SET CaseXref='{}', BeneficiaryId='{}', SourceCreatedDate='{}', CasePetitionName='{}', CaseType='{}', CaseDescription='{}', CaseFiledDate='{}', ReceiptNumber='{}', ReceiptStatus='{}', RFEAuditReceivedDate='{}', RFEAuditDueDate='{}', RFEAuditSubmittedDate='{}', PrimaryCaseStatus='{}', SecondaryCaseStatus='{}', CaseComments='{}', LastStepCompleted='{}', LastStepCompletedDate='{}', NextStepAction='{}', NextStepActionDueDate='{}', PriorityDate='{}', PriorityCategory='{}', PriorityCountry='{}', CaseApprovedDate='{}', CaseValidFromDate='{}', CaseExpirationDate='{}', CaseClosedDate='{}', CaseDeniedDate='{}', CaseWithdrawnDate='{}', CasePrimaryAttorney='{}', CaseReviewingAttorney='{}', CasePrimaryCaseManager='{}', PetitionXref='{}', from_name='{}' WHERE CaseId='{}'".format(case_xref, beneficiary_id, case_creation_date, case_petition_name, case_type, case_description, case_filed_date, case_receipt_number, case_receipt_status, rfe_audit_received_date, rfe_audit_due_date, rfe_audit_submitted_date, primary_case_status, secondary_case_status, case_comments, case_last_step_completed, case_last_step_completed_date, case_next_step_to_be_completed, case_next_step_to_be_completed_date, case_priority_date, case_priority_category, case_priority_country, case_approved_date, case_valid_from, case_valid_to, case_closed_date, case_denied_date, case_withdrawn_date, case_primary_attorney, case_reviewing_attorney, case_primary_case_manager, petition_xref, from_name, case_id))
                        cursor.commit()



                
def generate_case_report(todate=''):
    result_filepath = 'Processed Reports Dashboard/GT_Immigration Practice Management Dashboard Data'+'_'+str(todate)+'.xlsx'
    process_report(result_filepath)
    print('Report Generated')
                
                


            
def process_report(result_filepath):

    
    #Tab 1 - Case Initiations
    wb_r = Workbook(result_filepath)
    

    bold = wb_r.add_format({'bold': True})
    
    formatdict = {'num_format': 'dd-mmm-yyyy'}
    fmt = wb_r.add_format()
    fmt.set_text_wrap()

    fmt_d = wb_r.add_format(formatdict)
    fmt_d.set_text_wrap()
    header_cell_format = wb_r.add_format()
    header_cell_format.set_text_wrap()
    # This is optional when using a solid fill.
    #header_cell_format.set_pattern(1) 
    header_cell_format.set_font_size(12) 

    header_cell_format2 = wb_r.add_format({'bold': True})
    header_cell_format2.set_text_wrap()
    # This is optional when using a solid fill.
    header_cell_format2.set_pattern(1)
    header_cell_format2.set_bg_color('#cce6ff')
    header_cell_format2.set_font_color('black')
    header_cell_format2.set_font_size(10) 
    
    ###################################### Tab 0 Header #############################################
    ws_r = wb_r.add_worksheet("Combined")
    ws_r.hide_gridlines(0)
    
    headers_c = ['Beneficiary Id', 'Organization Name', 'Petitioner Name','Petitioner of Primary Beneficiary','Beneficiary Type (Employee / Dependent)', 'Beneficiary Record Creation Date', 'Beneficiary Record Status', 'Beneficiary Record Inactivation Date', 'Beneficiary Last Name', 'Beneficiary First Name', 'Primary Beneficiary Id', 'Primary Beneficiary Last Name', 'Primary Beneficiary First Name', 'Relation',  'Immigration Status (I-94)', 'Immigration Status Expiration Date (I-94)', 'I-797 Approved Date', 'I-797 Status', 'I-797 Expiration Date', 'Final NIV (Maxout) Date', 'Maxout Date Applicability and Note', 'Case Id', 'Case Creation Date', 'Case Petition Name', 'Case Type', 'Case Description', 'Case Filed Date', 'Case Receipt Number', 'Case Receipt Status', 'RFE/Audit Received Date', 'RFE/Audit Response Due Date', 'RFE/Audit Response Submitted Date', 'Primary Case Status', 'Secondary Case Status', 'Case Comments', 'Case Last Step Completed', 'Case Last Step Completed Date', 'Case Next Step To Be Completed', 'Case Next Step To Be Completed Date', 'Case Priority Date', 'Case Priority Category', 'Case Priority Country', 'Case Approved Date', 'Case Valid From', 'Case Valid To', 'Case Closed Date', 'Case Denied Date', 'Case Withdrawn Date', 'Case Primary Attorney', 'Reviewing Attorney', 'Case Manager']

   
    headers_table_c = ['BeneficiaryXref', 'OrganizationName', 'PetitionerName','Primary_Petitioner','BeneficiaryType', 'SourceCreatedDate', 'BeneficiaryRecordStatus', 'InactiveDate', 'LastName', 'FirstName', 'PrimaryBeneficiaryXref', 'PrimaryBeneficiaryLastName', 'PrimaryBeneficiaryFirstName', 'RelationType', 'ImmigrationStatus', 'ImmigrationStatusExpirationDate', 'MostRecentI797IssueApprovalDate', 'MostRecentI797Status', 'I797ExpirationDate', 'FinalNivDate', 'MaxOutDateNote', 'CaseXref','CaseSourceCreatedDate', 'CasePetitionName', 'CaseType', 'CaseDescription', 'CaseFiledDate','ReceiptNumber','ReceiptStatus','RFEAuditReceivedDate','RFEAuditDueDate','RFEAuditSubmittedDate', 'PrimaryCaseStatus','SecondaryCaseStatus','CaseComments','LastStepCompleted','LastStepCompletedDate', 'NextStepAction', 'NextStepActionDueDate', 'PriorityDate','PriorityCategory','PriorityCountry','CaseApprovedDate','CaseValidFromDate','CaseExpirationDate','CaseClosedDate','CaseDeniedDate','CaseWithdrawnDate', 'CasePrimaryAttorney', 'CaseReviewingAttorney', 'CasePrimaryCaseManager']
    
   

    header_names_c = [{'header': x} for x in headers_c]

    results_active_c = cursor.execute("SELECT b.*,c.CaseXref,c.SourceCreatedDate as CaseSourceCreatedDate,c.CasePetitionName,\
        c.CaseType, c.CaseDescription, c.CaseFiledDate, c.ReceiptNumber,c.ReceiptStatus, c.RFEAuditReceivedDate,c.RFEAuditDueDate, c.RFEAuditSubmittedDate, \
        c.PrimaryCaseStatus, c.SecondaryCaseStatus, c.CaseComments, c.LastStepCompleted, c.LastStepCompletedDate, c.NextStepAction, c.NextStepActionDueDate, c.PriorityDate, c.PriorityCategory, c.PriorityCountry, c.CaseApprovedDate, c.CaseValidFromDate, c.CaseExpirationDate, c.CaseClosedDate, c.CaseDeniedDate, c.CaseWithdrawnDate, c.CasePrimaryCaseManager,\
        c.CasePrimaryAttorney, c.CaseReviewingAttorney,\
        o.OrganizationName,p.PetitionerName,p2.PetitionerName as Primary_Petitioner,\
        CASE WHEN b.IsActive=1 THEN 'Active' ELSE 'Inactive' END as BeneficiaryRecordStatus, \
        e.EmployeeId\
        FROM dbo.[Case] as c  \
        LEFT JOIN  dbo.Beneficiary as b on c.BeneficiaryId=b.BeneficiaryId\
        LEFT JOIN dbo.Petitioner as p on b.PetitionerId=p.PetitionerId\
        LEFT JOIN dbo.Organization as o on p.OrganizationId=o.OrganizationId\
        LEFT JOIN dbo.Beneficiary as b2 on b.PrimaryBeneficiaryXref=b2.BeneficiaryXref\
        LEFT JOIN dbo.Petitioner as p2 on b2.PetitionerId=p2.PetitionerId\
        LEFT JOIN dbo.Organization as o2 on p2.OrganizationId=o2.OrganizationId\
        LEFT JOIN dbo.BeneficiaryEmployment as e on e.BeneficiaryId=b.BeneficiaryId\
        where (c.SourceCreatedDate >= '2020-01-01'  \
        or c.CaseFiledDate >= '2020-01-01' \
        or (b.IsActive = '1'  and c.PrimaryCaseStatus='open'))\
        ORDER BY b.is_primary_beneficiary DESC, b.LastName ASC  ").fetchall()
    count = len(results_active_c)
    if count == 0:
        count = count + 1
    ws_r.add_table('A1:AY'+str(count+1),{'columns': header_names_c, 'style': 'Table Style Medium 2'})
    for idx, hdr in enumerate(headers_c):
        col = headers_c.index(hdr)
        ##print(col, hdr)
        ws_r.write(0, col, hdr, header_cell_format)
        ws_r.set_column(col, col, 14, fmt)


    
    ###################################### Tab 1 Header #############################################
    ws_r = wb_r.add_worksheet("Case Initiations")
    ws_r.hide_gridlines(0)
    
    headers = ['Beneficiary Id', 'Organization Name', 'Petitioner Name','Petitioner of Primary Beneficiary','Beneficiary Type (Employee / Dependent)', 'Beneficiary Record Creation Date', 'Beneficiary Record Status', 'Beneficiary Record Inactivation Date', 'Beneficiary Last Name', 'Beneficiary First Name', 'Primary Beneficiary Id', 'Primary Beneficiary Last Name', 'Primary Beneficiary First Name', 'Relation',  'Immigration Status (I-94)', 'Immigration Status Expiration Date (I-94)', 'I-797 Approved Date', 'I-797 Status', 'I-797 Expiration Date', 'Final NIV (Maxout) Date', 'Maxout Date Applicability and Note', 'Case Id', 'Case Creation Date', 'Case Petition Name', 'Case Type', 'Case Description', 'Case Filed Date', 'Case Receipt Number', 'Case Receipt Status', 'RFE/Audit Received Date', 'RFE/Audit Response Due Date', 'RFE/Audit Response Submitted Date', 'Primary Case Status', 'Secondary Case Status', 'Case Comments', 'Case Last Step Completed', 'Case Last Step Completed Date', 'Case Next Step To Be Completed', 'Case Next Step To Be Completed Date', 'Case Priority Date', 'Case Priority Category', 'Case Priority Country', 'Case Approved Date', 'Case Valid From', 'Case Valid To', 'Case Closed Date', 'Case Denied Date', 'Case Withdrawn Date', 'Case Primary Attorney', 'Reviewing Attorney', 'Case Manager']

   
    headers_table = ['BeneficiaryXref', 'OrganizationName', 'PetitionerName','Primary_Petitioner','BeneficiaryType', 'SourceCreatedDate', 'BeneficiaryRecordStatus', 'InactiveDate', 'LastName', 'FirstName', 'PrimaryBeneficiaryXref', 'PrimaryBeneficiaryLastName', 'PrimaryBeneficiaryFirstName', 'RelationType', 'ImmigrationStatus', 'ImmigrationStatusExpirationDate', 'MostRecentI797IssueApprovalDate', 'MostRecentI797Status', 'I797ExpirationDate', 'FinalNivDate', 'MaxOutDateNote', 'CaseXref','CaseSourceCreatedDate', 'CasePetitionName', 'CaseType', 'CaseDescription', 'CaseFiledDate','ReceiptNumber','ReceiptStatus','RFEAuditReceivedDate','RFEAuditDueDate','RFEAuditSubmittedDate', 'PrimaryCaseStatus','SecondaryCaseStatus','CaseComments','LastStepCompleted','LastStepCompletedDate', 'NextStepAction', 'NextStepActionDueDate', 'PriorityDate','PriorityCategory','PriorityCountry','CaseApprovedDate','CaseValidFromDate','CaseExpirationDate','CaseClosedDate','CaseDeniedDate','CaseWithdrawnDate', 'CasePrimaryAttorney', 'CaseReviewingAttorney', 'CasePrimaryCaseManager']
    
   

    header_names = [{'header': x} for x in headers]

    results_active = cursor.execute("SELECT b.*,c.CaseXref,c.SourceCreatedDate as CaseSourceCreatedDate,c.CasePetitionName,\
        c.CaseType, c.CaseDescription, c.CaseFiledDate, c.ReceiptNumber,c.ReceiptStatus, c.RFEAuditReceivedDate,c.RFEAuditDueDate, c.RFEAuditSubmittedDate, \
        c.PrimaryCaseStatus, c.SecondaryCaseStatus, c.CaseComments, c.LastStepCompleted, c.LastStepCompletedDate, c.NextStepAction, c.NextStepActionDueDate, c.PriorityDate, c.PriorityCategory, c.PriorityCountry, c.CaseApprovedDate, c.CaseValidFromDate, c.CaseExpirationDate, c.CaseClosedDate, c.CaseDeniedDate, c.CaseWithdrawnDate, c.CasePrimaryCaseManager,\
        c.CasePrimaryAttorney, c.CaseReviewingAttorney,\
        o.OrganizationName,p.PetitionerName,p2.PetitionerName as Primary_Petitioner,\
        CASE WHEN b.IsActive=1 THEN 'Active' ELSE 'Inactive' END as BeneficiaryRecordStatus, \
        e.EmployeeId\
        FROM dbo.[Case] as c  \
        LEFT JOIN  dbo.Beneficiary as b on c.BeneficiaryId=b.BeneficiaryId\
        LEFT JOIN dbo.Petitioner as p on b.PetitionerId=p.PetitionerId\
        LEFT JOIN dbo.Organization as o on p.OrganizationId=o.OrganizationId\
        LEFT JOIN dbo.Beneficiary as b2 on b.PrimaryBeneficiaryXref=b2.BeneficiaryXref\
        LEFT JOIN dbo.Petitioner as p2 on b2.PetitionerId=p2.PetitionerId\
        LEFT JOIN dbo.Organization as o2 on p2.OrganizationId=o2.OrganizationId\
        LEFT JOIN dbo.BeneficiaryEmployment as e on e.BeneficiaryId=b.BeneficiaryId\
        where c.SourceCreatedDate >= '2020-01-01'  \
        ORDER BY b.is_primary_beneficiary DESC, b.LastName ASC  ").fetchall()
    count = len(results_active)
    if count == 0:
        count = count + 1
    ws_r.add_table('A1:AY'+str(count+1),{'columns': header_names, 'style': 'Table Style Medium 2'})
    for idx, hdr in enumerate(headers):
        col = headers.index(hdr)
        ##print(col, hdr)
        ws_r.write(0, col, hdr, header_cell_format)
        ws_r.set_column(col, col, 14, fmt)

    

    ###################################### Tab 2 Header #############################################
    ws_r2 = wb_r.add_worksheet("Case Filings")
    ws_r2.hide_gridlines(0)

    
    headers2 = ['Beneficiary Id', 'Organization Name', 'Petitioner Name','Petitioner of Primary Beneficiary','Beneficiary Type (Employee / Dependent)', 'Beneficiary Record Creation Date', 'Beneficiary Record Status', 'Beneficiary Record Inactivation Date', 'Beneficiary Last Name', 'Beneficiary First Name', 'Primary Beneficiary Id', 'Primary Beneficiary Last Name', 'Primary Beneficiary First Name', 'Relation',  'Immigration Status (I-94)', 'Immigration Status Expiration Date (I-94)', 'I-797 Approved Date', 'I-797 Status', 'I-797 Expiration Date', 'Final NIV (Maxout) Date', 'Maxout Date Applicability and Note', 'Case Id', 'Case Creation Date', 'Case Petition Name', 'Case Type', 'Case Description', 'Case Filed Date', 'Case Receipt Number', 'Case Receipt Status', 'RFE/Audit Received Date', 'RFE/Audit Response Due Date', 'RFE/Audit Response Submitted Date', 'Primary Case Status', 'Secondary Case Status', 'Case Comments', 'Case Last Step Completed', 'Case Last Step Completed Date', 'Case Next Step To Be Completed', 'Case Next Step To Be Completed Date', 'Case Priority Date', 'Case Priority Category', 'Case Priority Country', 'Case Approved Date', 'Case Valid From', 'Case Valid To', 'Case Closed Date', 'Case Denied Date', 'Case Withdrawn Date', 'Case Primary Attorney', 'Reviewing Attorney', 'Case Manager']

   
    headers_table2 = ['BeneficiaryXref', 'OrganizationName', 'PetitionerName','Primary_Petitioner','BeneficiaryType', 'SourceCreatedDate', 'BeneficiaryRecordStatus', 'InactiveDate', 'LastName', 'FirstName', 'PrimaryBeneficiaryXref', 'PrimaryBeneficiaryLastName', 'PrimaryBeneficiaryFirstName', 'RelationType', 'ImmigrationStatus', 'ImmigrationStatusExpirationDate', 'MostRecentI797IssueApprovalDate', 'MostRecentI797Status', 'I797ExpirationDate', 'FinalNivDate', 'MaxOutDateNote', 'CaseXref','CaseSourceCreatedDate', 'CasePetitionName', 'CaseType', 'CaseDescription', 'CaseFiledDate','ReceiptNumber','ReceiptStatus','RFEAuditReceivedDate','RFEAuditDueDate','RFEAuditSubmittedDate', 'PrimaryCaseStatus','SecondaryCaseStatus','CaseComments','LastStepCompleted','LastStepCompletedDate', 'NextStepAction', 'NextStepActionDueDate', 'PriorityDate','PriorityCategory','PriorityCountry','CaseApprovedDate','CaseValidFromDate','CaseExpirationDate','CaseClosedDate','CaseDeniedDate','CaseWithdrawnDate', 'CasePrimaryAttorney', 'CaseReviewingAttorney', 'CasePrimaryCaseManager']
    
   
    column_till = 'AY'
    header_names2 = [{'header': x} for x in headers2]

    results_active2 = cursor.execute("SELECT b.*,c.CaseXref,c.SourceCreatedDate as CaseSourceCreatedDate,c.CasePetitionName,\
        c.CaseType, c.CaseDescription, c.CaseFiledDate, c.ReceiptNumber,c.ReceiptStatus, c.RFEAuditReceivedDate,c.RFEAuditDueDate, c.RFEAuditSubmittedDate, \
        c.PrimaryCaseStatus, c.SecondaryCaseStatus, c.CaseComments, c.LastStepCompleted, c.LastStepCompletedDate, c.NextStepAction, c.NextStepActionDueDate, c.PriorityDate, c.PriorityCategory, c.PriorityCountry, c.CaseApprovedDate, c.CaseValidFromDate, c.CaseExpirationDate, c.CaseClosedDate, c.CaseDeniedDate, c.CaseWithdrawnDate, c.CasePrimaryCaseManager,\
        c.CasePrimaryAttorney, c.CaseReviewingAttorney,\
        o.OrganizationName,p.PetitionerName,p2.PetitionerName as Primary_Petitioner,\
        CASE WHEN b.IsActive=1 THEN 'Active' ELSE 'Inactive' END as BeneficiaryRecordStatus, \
        e.EmployeeId\
        FROM dbo.[Case] as c  \
        LEFT JOIN  dbo.Beneficiary as b on c.BeneficiaryId=b.BeneficiaryId\
        LEFT JOIN dbo.Petitioner as p on b.PetitionerId=p.PetitionerId\
        LEFT JOIN dbo.Organization as o on p.OrganizationId=o.OrganizationId\
        LEFT JOIN dbo.Beneficiary as b2 on b.PrimaryBeneficiaryXref=b2.BeneficiaryXref\
        LEFT JOIN dbo.Petitioner as p2 on b2.PetitionerId=p2.PetitionerId\
        LEFT JOIN dbo.Organization as o2 on p2.OrganizationId=o2.OrganizationId\
        LEFT JOIN dbo.BeneficiaryEmployment as e on e.BeneficiaryId=b.BeneficiaryId\
        where c.CaseFiledDate >= '2020-01-01' \
        ORDER BY b.is_primary_beneficiary DESC, b.LastName ASC  ").fetchall()

        
    count2 = len(results_active2)
    if count2 == 0 : 
        count2 = 1
    ws_r2.add_table('A1:'+str(column_till)+str(count2+1),{'columns': header_names2, 'style': 'Table Style Medium 2'})
    for idx, hdr in enumerate(headers2):
        col = headers2.index(hdr)
        ##print(col, hdr)
        ws_r2.write(0, col, hdr, header_cell_format)
        ws_r2.set_column(col, col, 14, fmt)
    
    
    ###################################### Tab 3 Header #############################################
    ws_r3 = wb_r.add_worksheet("Open Cases")
    ws_r3.hide_gridlines(0)

    headers3 = ['Beneficiary Id', 'Organization Name', 'Petitioner Name','Petitioner of Primary Beneficiary','Beneficiary Type (Employee / Dependent)', 'Beneficiary Record Creation Date', 'Beneficiary Record Status', 'Beneficiary Record Inactivation Date', 'Beneficiary Last Name', 'Beneficiary First Name', 'Primary Beneficiary Id', 'Primary Beneficiary Last Name', 'Primary Beneficiary First Name', 'Relation',  'Immigration Status (I-94)', 'Immigration Status Expiration Date (I-94)', 'I-797 Approved Date', 'I-797 Status', 'I-797 Expiration Date', 'Final NIV (Maxout) Date', 'Maxout Date Applicability and Note', 'Case Id', 'Case Creation Date', 'Case Petition Name', 'Case Type', 'Case Description', 'Case Filed Date', 'Case Receipt Number', 'Case Receipt Status', 'RFE/Audit Received Date', 'RFE/Audit Response Due Date', 'RFE/Audit Response Submitted Date', 'Primary Case Status', 'Secondary Case Status', 'Case Comments', 'Case Last Step Completed', 'Case Last Step Completed Date', 'Case Next Step To Be Completed', 'Case Next Step To Be Completed Date', 'Case Priority Date', 'Case Priority Category', 'Case Priority Country', 'Case Approved Date', 'Case Valid From', 'Case Valid To', 'Case Closed Date', 'Case Denied Date', 'Case Withdrawn Date', 'Case Primary Attorney', 'Reviewing Attorney', 'Case Manager']

   
    headers_table3 = ['BeneficiaryXref', 'OrganizationName', 'PetitionerName','Primary_Petitioner','BeneficiaryType', 'SourceCreatedDate', 'BeneficiaryRecordStatus', 'InactiveDate', 'LastName', 'FirstName', 'PrimaryBeneficiaryXref', 'PrimaryBeneficiaryLastName', 'PrimaryBeneficiaryFirstName', 'RelationType', 'ImmigrationStatus', 'ImmigrationStatusExpirationDate', 'MostRecentI797IssueApprovalDate', 'MostRecentI797Status', 'I797ExpirationDate', 'FinalNivDate', 'MaxOutDateNote', 'CaseXref','CaseSourceCreatedDate', 'CasePetitionName', 'CaseType', 'CaseDescription', 'CaseFiledDate','ReceiptNumber','ReceiptStatus','RFEAuditReceivedDate','RFEAuditDueDate','RFEAuditSubmittedDate', 'PrimaryCaseStatus','SecondaryCaseStatus','CaseComments','LastStepCompleted','LastStepCompletedDate', 'NextStepAction', 'NextStepActionDueDate', 'PriorityDate','PriorityCategory','PriorityCountry','CaseApprovedDate','CaseValidFromDate','CaseExpirationDate','CaseClosedDate','CaseDeniedDate','CaseWithdrawnDate', 'CasePrimaryAttorney', 'CaseReviewingAttorney', 'CasePrimaryCaseManager']

    column_till = 'AY'

    header_names3 = [{'header': x} for x in headers3]

    results_active3 = cursor.execute("SELECT b.*,c.CaseXref,c.SourceCreatedDate as CaseSourceCreatedDate,c.CasePetitionName,\
        c.CaseType, c.CaseDescription, c.CaseFiledDate, c.ReceiptNumber,c.ReceiptStatus, c.RFEAuditReceivedDate,c.RFEAuditDueDate, c.RFEAuditSubmittedDate, \
        c.PrimaryCaseStatus, c.SecondaryCaseStatus, c.CaseComments, c.LastStepCompleted, c.LastStepCompletedDate, c.NextStepAction, c.NextStepActionDueDate, c.PriorityDate, c.PriorityCategory, c.PriorityCountry, c.CaseApprovedDate, c.CaseValidFromDate, c.CaseExpirationDate, c.CaseClosedDate, c.CaseDeniedDate, c.CaseWithdrawnDate, c.CasePrimaryCaseManager,\
        c.CasePrimaryAttorney, c.CaseReviewingAttorney,\
        o.OrganizationName,p.PetitionerName,p2.PetitionerName as Primary_Petitioner,\
        CASE WHEN b.IsActive=1 THEN 'Active' ELSE 'Inactive' END as BeneficiaryRecordStatus, \
        e.EmployeeId\
        FROM dbo.[Case] as c  \
        LEFT JOIN  dbo.Beneficiary as b on c.BeneficiaryId=b.BeneficiaryId\
        LEFT JOIN dbo.Petitioner as p on b.PetitionerId=p.PetitionerId\
        LEFT JOIN dbo.Organization as o on p.OrganizationId=o.OrganizationId\
        LEFT JOIN dbo.Beneficiary as b2 on b.PrimaryBeneficiaryXref=b2.BeneficiaryXref\
        LEFT JOIN dbo.Petitioner as p2 on b2.PetitionerId=p2.PetitionerId\
        LEFT JOIN dbo.Organization as o2 on p2.OrganizationId=o2.OrganizationId\
        LEFT JOIN dbo.BeneficiaryEmployment as e on e.BeneficiaryId=b.BeneficiaryId\
        where b.IsActive = '1'  and c.PrimaryCaseStatus='open' \
        ORDER BY b.is_primary_beneficiary DESC, b.LastName ASC  ").fetchall()
    count3 = len(results_active3)
    if count3 == 0 : 
        count3 = 1
    ws_r3.add_table('A1:'+str(column_till)+str(count3+1),{'columns': header_names3, 'style': 'Table Style Medium 2'})
    for idx, hdr in enumerate(headers3):
        col = headers3.index(hdr)
        ##print(col, hdr)
        ws_r3.write(0, col, hdr, header_cell_format)
        ws_r3.set_column(col, col, 14, fmt)
    
    wb_r.close()  
    
    ###################################### Tab 0 - Combined #############################################
    wb_pyxl = load_workbook(result_filepath)  
    wb_pyxl.active = 0 #active first sheet
    sheet = wb_pyxl.active 
    
    sheet.freeze_panes = 'B2'
    color_format  = PatternFill(start_color="FFFF66",end_color="FFFF66", fill_type = "solid") 
    for hdr in headers_table_c:
        col = headers_table_c.index(hdr)
        sheet.cell(row=1, column=col+1).alignment=Alignment(wrap_text=True, horizontal="justify", vertical="justify")

    for _key, s in enumerate(results_active_c):
        num = _key+1
        is_primary_beneficiary = getattr(s, 'is_primary_beneficiary')
        for hdr in headers_table_c:
            col = headers_table_c.index(hdr)
            
            if hdr:
                value_obj = getattr(s, hdr)
            else: 
                value_obj = ''
            
            ##print(int(num)+1, int(col)+1, hdr, value_obj)
            

            if hdr == "SourceCreatedDate" or hdr == "InactiveDate" or hdr == "ImmigrationStatusExpirationDate" or hdr == "I797ExpirationDate" or hdr == "FinalNivDate" or hdr == "VisaPedDate" or hdr == "EadExpirationDate" or hdr == "AdvanceParoleExpirationDate" or hdr == "EadApExpirationDate" or hdr == "Ds2019ValidFromDate"  or hdr == "Ds2019ExpirationDate"  or hdr == "ReEntryPermitExpirationDate"  or hdr == "GreenCardExpirationDate"  or hdr == "MostRecentPassportExpirationDate"  or hdr == "VisaExpirationDate"  or hdr == "HireDate" or hdr == "Priority1Date"  or hdr == "Priority2Date"  or hdr == "Priority3Date"  or hdr == "Priority4Date"  or hdr == "Priority5Date"  or hdr == "NextStepActionDueDate" or hdr == "MostRecentI797IssueApprovalDate" or hdr == "CaseSourceCreatedDate" or hdr == "CaseFiledDate" or hdr == "RFEAuditReceivedDate" or hdr == "RFEAuditDueDate" or hdr == "RFEAuditSubmittedDate"  or hdr == "LastStepCompletedDate"  or hdr == "NextStepActionDueDate"  or hdr == "PriorityDate"  or hdr == "CaseApprovedDate"  or hdr == "CaseValidFromDate"  or hdr == "CaseExpirationDate" or hdr == "CaseClosedDate"  or hdr == "CaseDeniedDate"  or hdr == "CaseWithdrawnDate":
                ##print(value_obj)
                if hdr == "ImmigrationStatusExpirationDate":
                    if value_obj.strip() != "D/S":
                        if 'D/S' in value_obj.strip():
                            split1 = (value_obj.strip()).split('(D/S)')
                            value_obj = change_display_format(str(split1[0]).replace('00:00:00', ''))
                            value_obj = str(value_obj)+' (D/S)'
                        else:
                            value_obj = change_display_format(str(value_obj).replace('00:00:00', ''))
                        
                else:        
                    if str(value_obj) == '1900-01-01 00:00:00':
                        value_obj = ''
                    else:
                        value_obj = change_display_format(str(value_obj).replace('00:00:00', ''))

                if value_obj is None or value_obj == "None":
                    value_obj = ''

                sheet.cell(row=int(num)+1, column=col+1).value = str(value_obj) 
                sheet.cell(row=int(num)+1, column=col+1).alignment=Alignment(wrap_text=True, horizontal="justify", vertical="justify")
                sheet.cell(row=int(num)+1, column=col+1).number_format = 'd-mmm-yy'
                sheet.cell(row=int(num)+1, column=col+1).font= Font(name = 'Calibri (Body)', size= 11)
                sheet.cell(row=int(num)+1, column=col+1).border= Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
                
                pass
            else:
                if hdr == "OrganizationName":
                    if is_primary_beneficiary == 0:
                        value_obj = ''
                elif hdr == "PetitionerName":
                    if is_primary_beneficiary == 0:
                        value_obj = 'Individual'
                elif hdr == "BeneficiaryType":
                    if value_obj.casefold() == 'Yes'.casefold():
                        value_obj = "Employee"
                    elif value_obj.casefold() == 'No'.casefold():
                        value_obj = "Dependent"


                if value_obj is None or value_obj == "None":
                    value_obj = ''

                sheet.cell(row=int(num)+1, column = int(col)+1).value = str(value_obj) 
                sheet.cell(row=int(num)+1, column=col+1).alignment=Alignment(wrap_text=True, horizontal="justify", vertical="justify")
                sheet.cell(row=int(num)+1, column=col+1).font= Font(name = 'Calibri (Body)', size= 11)
                sheet.cell(row=int(num)+1, column=col+1).border= Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                
                pass
            
        #return False
    wb_pyxl.save(result_filepath)
    
    ###################################### Tab 1 - Data #############################################
    wb_pyxl = load_workbook(result_filepath)  
    wb_pyxl.active = 1 #active first sheet
    sheet = wb_pyxl.active 
    
    sheet.freeze_panes = 'B2'
    color_format  = PatternFill(start_color="FFFF66",end_color="FFFF66", fill_type = "solid") 
    for hdr in headers_table:
        col = headers_table.index(hdr)
        sheet.cell(row=1, column=col+1).alignment=Alignment(wrap_text=True, horizontal="justify", vertical="justify")

    for _key, s in enumerate(results_active):
        num = _key+1
        is_primary_beneficiary = getattr(s, 'is_primary_beneficiary')
        for hdr in headers_table:
            col = headers_table.index(hdr)
            
            if hdr:
                value_obj = getattr(s, hdr)
            else: 
                value_obj = ''
            
            ##print(int(num)+1, int(col)+1, hdr, value_obj)
            

            if hdr == "SourceCreatedDate" or hdr == "InactiveDate" or hdr == "ImmigrationStatusExpirationDate" or hdr == "I797ExpirationDate" or hdr == "FinalNivDate" or hdr == "VisaPedDate" or hdr == "EadExpirationDate" or hdr == "AdvanceParoleExpirationDate" or hdr == "EadApExpirationDate" or hdr == "Ds2019ValidFromDate"  or hdr == "Ds2019ExpirationDate"  or hdr == "ReEntryPermitExpirationDate"  or hdr == "GreenCardExpirationDate"  or hdr == "MostRecentPassportExpirationDate"  or hdr == "VisaExpirationDate"  or hdr == "HireDate" or hdr == "Priority1Date"  or hdr == "Priority2Date"  or hdr == "Priority3Date"  or hdr == "Priority4Date"  or hdr == "Priority5Date"  or hdr == "NextStepActionDueDate" or hdr == "MostRecentI797IssueApprovalDate" or hdr == "CaseSourceCreatedDate" or hdr == "CaseFiledDate" or hdr == "RFEAuditReceivedDate" or hdr == "RFEAuditDueDate" or hdr == "RFEAuditSubmittedDate"  or hdr == "LastStepCompletedDate"  or hdr == "NextStepActionDueDate"  or hdr == "PriorityDate"  or hdr == "CaseApprovedDate"  or hdr == "CaseValidFromDate"  or hdr == "CaseExpirationDate" or hdr == "CaseClosedDate"  or hdr == "CaseDeniedDate"  or hdr == "CaseWithdrawnDate":
                ##print(value_obj)
                if hdr == "ImmigrationStatusExpirationDate":
                    if value_obj.strip() != "D/S":
                        if 'D/S' in value_obj.strip():
                            split1 = (value_obj.strip()).split('(D/S)')
                            value_obj = change_display_format(str(split1[0]).replace('00:00:00', ''))
                            value_obj = str(value_obj)+' (D/S)'
                        else:
                            value_obj = change_display_format(str(value_obj).replace('00:00:00', ''))
                        
                else:        
                    if str(value_obj) == '1900-01-01 00:00:00':
                        value_obj = ''
                    else:
                        value_obj = change_display_format(str(value_obj).replace('00:00:00', ''))

                if value_obj is None or value_obj == "None":
                    value_obj = ''

                sheet.cell(row=int(num)+1, column=col+1).value = str(value_obj) 
                sheet.cell(row=int(num)+1, column=col+1).alignment=Alignment(wrap_text=True, horizontal="justify", vertical="justify")
                sheet.cell(row=int(num)+1, column=col+1).number_format = 'd-mmm-yy'
                sheet.cell(row=int(num)+1, column=col+1).font= Font(name = 'Calibri (Body)', size= 11)
                sheet.cell(row=int(num)+1, column=col+1).border= Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
                
                pass
            else:
                if hdr == "OrganizationName":
                    if is_primary_beneficiary == 0:
                        value_obj = ''
                elif hdr == "PetitionerName":
                    if is_primary_beneficiary == 0:
                        value_obj = 'Individual'
                elif hdr == "BeneficiaryType":
                    if value_obj.casefold() == 'Yes'.casefold():
                        value_obj = "Employee"
                    elif value_obj.casefold() == 'No'.casefold():
                        value_obj = "Dependent"


                if value_obj is None or value_obj == "None":
                    value_obj = ''

                sheet.cell(row=int(num)+1, column = int(col)+1).value = str(value_obj) 
                sheet.cell(row=int(num)+1, column=col+1).alignment=Alignment(wrap_text=True, horizontal="justify", vertical="justify")
                sheet.cell(row=int(num)+1, column=col+1).font= Font(name = 'Calibri (Body)', size= 11)
                sheet.cell(row=int(num)+1, column=col+1).border= Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                
                pass
            
        #return False
    wb_pyxl.save(result_filepath)


      
    ###################################### Tab 2 - Data #############################################
    wb_pyxl = load_workbook(result_filepath)  
    wb_pyxl.active = 2 #active Second sheet
    sheet = wb_pyxl.active 
    sheet.freeze_panes = 'B2'
    color_format  = PatternFill(start_color="FFFF66",end_color="FFFF66", fill_type = "solid") 
    for hdr in headers_table:
        col = headers_table.index(hdr)
        sheet.cell(row=1, column=col+1).alignment=Alignment(wrap_text=True, horizontal="justify", vertical="justify")

    for _key, s in enumerate(results_active2):
        num = _key+1
        is_primary_beneficiary = getattr(s, 'is_primary_beneficiary')
        for hdr in headers_table2:
            col = headers_table2.index(hdr)
            
            if hdr:
                value_obj = getattr(s, hdr)
            else: 
                value_obj = ''

            
            ##print(int(num)+1, int(col)+1, hdr, value_obj)
            
            if hdr == "SourceCreatedDate" or hdr == "InactiveDate" or hdr == "ImmigrationStatusExpirationDate" or hdr == "I797ExpirationDate" or hdr == "FinalNivDate" or hdr == "VisaPedDate" or hdr == "EadExpirationDate" or hdr == "AdvanceParoleExpirationDate" or hdr == "EadApExpirationDate" or hdr == "Ds2019ValidFromDate"  or hdr == "Ds2019ExpirationDate"  or hdr == "ReEntryPermitExpirationDate"  or hdr == "GreenCardExpirationDate"  or hdr == "MostRecentPassportExpirationDate"  or hdr == "VisaExpirationDate"  or hdr == "HireDate" or hdr == "Priority1Date"  or hdr == "Priority2Date"  or hdr == "Priority3Date"  or hdr == "Priority4Date"  or hdr == "Priority5Date"  or hdr == "NextStepActionDueDate" or hdr == "MostRecentI797IssueApprovalDate" or hdr == "CaseSourceCreatedDate" or hdr == "CaseFiledDate" or hdr == "RFEAuditReceivedDate" or hdr == "RFEAuditDueDate" or hdr == "RFEAuditSubmittedDate"  or hdr == "LastStepCompletedDate"  or hdr == "NextStepActionDueDate"  or hdr == "PriorityDate"  or hdr == "CaseApprovedDate"  or hdr == "CaseValidFromDate"  or hdr == "CaseExpirationDate" or hdr == "CaseClosedDate"  or hdr == "CaseDeniedDate"  or hdr == "CaseWithdrawnDate":
                ##print(value_obj)
                if hdr == "ImmigrationStatusExpirationDate":
                    if value_obj.strip() != "D/S":
                        if 'D/S' in value_obj.strip():
                            split1 = (value_obj.strip()).split('(D/S)')
                            value_obj = change_display_format(str(split1[0]).replace('00:00:00', ''))
                            value_obj = str(value_obj)+' (D/S)'
                        else:
                            value_obj = change_display_format(str(value_obj).replace('00:00:00', ''))
                        
                else:        
                    if str(value_obj) == '1900-01-01 00:00:00':
                        value_obj = ''
                    else:
                        value_obj = change_display_format(str(value_obj).replace('00:00:00', ''))

                if value_obj is None or value_obj == "None":
                    value_obj = ''

                sheet.cell(row=int(num)+1, column=col+1).value = str(value_obj) 
                sheet.cell(row=int(num)+1, column=col+1).alignment=Alignment(wrap_text=True, horizontal="justify", vertical="justify")
                sheet.cell(row=int(num)+1, column=col+1).number_format = 'dd-mmm-yy'
                sheet.cell(row=int(num)+1, column=col+1).font= Font(name = 'Calibri (Body)', size= 11)
                sheet.cell(row=int(num)+1, column=col+1).border= Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
                
                pass
            else:
                if hdr == "OrganizationName":
                    if is_primary_beneficiary == 0:
                        value_obj = ''
                elif hdr == "PetitionerName":
                    if is_primary_beneficiary == 0:
                        value_obj = 'Individual'
                elif hdr == "BeneficiaryType":
                    if value_obj.casefold() == 'Yes'.casefold():
                        value_obj = "Employee"
                    elif value_obj.casefold() == 'No'.casefold():
                        value_obj = "Dependent"

                if value_obj is None or value_obj == "None":
                    value_obj = ''

                sheet.cell(row=int(num)+1, column = int(col)+1).value = str(value_obj) 
                sheet.cell(row=int(num)+1, column=col+1).alignment=Alignment(wrap_text=True, horizontal="justify", vertical="justify")
                sheet.cell(row=int(num)+1, column=col+1).font= Font(name = 'Calibri (Body)', size= 11)
                sheet.cell(row=int(num)+1, column=col+1).border= Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                
                pass
        
    wb_pyxl.save(result_filepath)
    ###################################### Tab 3 - Data #############################################
    wb_pyxl = load_workbook(result_filepath)  
    wb_pyxl.active = 3 #active Third sheet
    sheet = wb_pyxl.active 
    sheet.freeze_panes = 'B2'
    color_format  = PatternFill(start_color="FFFF66",end_color="FFFF66", fill_type = "solid") 
    for hdr in headers_table:
        col = headers_table.index(hdr)
        sheet.cell(row=1, column=col+1).alignment=Alignment(wrap_text=True, horizontal="justify", vertical="justify")

    for _key, s in enumerate(results_active3):
        num = _key+1
        is_primary_beneficiary = getattr(s, 'is_primary_beneficiary')
        for hdr in headers_table3:
            col = headers_table3.index(hdr)
            
            if hdr:
                value_obj = getattr(s, hdr)
            else: 
                value_obj = ''
            
            ##print(int(num)+1, int(col)+1, hdr, value_obj)
            
            if hdr == "SourceCreatedDate" or hdr == "InactiveDate" or hdr == "ImmigrationStatusExpirationDate" or hdr == "I797ExpirationDate" or hdr == "FinalNivDate" or hdr == "VisaPedDate" or hdr == "EadExpirationDate" or hdr == "AdvanceParoleExpirationDate" or hdr == "EadApExpirationDate" or hdr == "Ds2019ValidFromDate"  or hdr == "Ds2019ExpirationDate"  or hdr == "ReEntryPermitExpirationDate"  or hdr == "GreenCardExpirationDate"  or hdr == "MostRecentPassportExpirationDate"  or hdr == "VisaExpirationDate"  or hdr == "HireDate" or hdr == "Priority1Date"  or hdr == "Priority2Date"  or hdr == "Priority3Date"  or hdr == "Priority4Date"  or hdr == "Priority5Date"  or hdr == "NextStepActionDueDate" or hdr == "MostRecentI797IssueApprovalDate" or hdr == "CaseSourceCreatedDate" or hdr == "CaseFiledDate" or hdr == "RFEAuditReceivedDate" or hdr == "RFEAuditDueDate" or hdr == "RFEAuditSubmittedDate"  or hdr == "LastStepCompletedDate"  or hdr == "NextStepActionDueDate"  or hdr == "PriorityDate"  or hdr == "CaseApprovedDate"  or hdr == "CaseValidFromDate"  or hdr == "CaseExpirationDate" or hdr == "CaseClosedDate"  or hdr == "CaseDeniedDate"  or hdr == "CaseWithdrawnDate":
                ##print(value_obj)
                if hdr == "ImmigrationStatusExpirationDate":
                    if value_obj.strip() != "D/S":
                        if 'D/S' in value_obj.strip():
                            split1 = (value_obj.strip()).split('(D/S)')
                            value_obj = change_display_format(str(split1[0]).replace('00:00:00', ''))
                            value_obj = str(value_obj)+' (D/S)'
                        else:
                            value_obj = change_display_format(str(value_obj).replace('00:00:00', ''))
                        
                else:        
                    if str(value_obj) == '1900-01-01 00:00:00':
                        value_obj = ''
                    else:
                        value_obj = change_display_format(str(value_obj).replace('00:00:00', ''))

                if value_obj is None or value_obj == "None":
                    value_obj = ''

                sheet.cell(row=int(num)+1, column=col+1).value = str(value_obj) 
                sheet.cell(row=int(num)+1, column=col+1).alignment=Alignment(wrap_text=True, horizontal="justify", vertical="justify")
                sheet.cell(row=int(num)+1, column=col+1).number_format = 'dd-mmm-yy'
                sheet.cell(row=int(num)+1, column=col+1).font= Font(name = 'Calibri (Body)', size= 11)
                sheet.cell(row=int(num)+1, column=col+1).border= Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
                
                pass
            else:
                if hdr == "OrganizationName":
                    if is_primary_beneficiary == 0:
                        value_obj = ''
                elif hdr == "PetitionerName":
                    if is_primary_beneficiary == 0:
                        value_obj = 'Individual'
                elif hdr == "BeneficiaryType":
                    if value_obj.casefold() == 'Yes'.casefold():
                        value_obj = "Employee"
                    elif value_obj.casefold() == 'No'.casefold():
                        value_obj = "Dependent"

                if value_obj is None or value_obj == "None":
                    value_obj = ''

                sheet.cell(row=int(num)+1, column = int(col)+1).value = str(value_obj) 
                sheet.cell(row=int(num)+1, column=col+1).alignment=Alignment(wrap_text=True, horizontal="justify", vertical="justify")
                sheet.cell(row=int(num)+1, column=col+1).font= Font(name = 'Calibri (Body)', size= 11)
                sheet.cell(row=int(num)+1, column=col+1).border= Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                
                pass
    
    
    wb_pyxl.active = 0 #active First sheet
    wb_pyxl.save(result_filepath)

    
    pass
        



if __name__ == '__main__':
    start()
    
    
    

