from datetime import datetime, date
import pandas as pd
import glob
# from xlsxwriter import Workbook
import pyodbc
import re

from openpyxl import formatting, styles, Workbook as openpyxl_workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, Color
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.styles.fills import Fill
from openpyxl.formatting.rule import CellIsRule, Rule
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter



conn = pyodbc.connect('Driver={ODBC Driver 17 for SQL Server};'
                      'Server=DESKTOP-GLMUIDH\SQLEXPRESS;'
                      'Database=VISA_BULLETIN_IMPORT;'
                      'Trusted_Connection=yes;')

cursor = conn.cursor()


def change_format(date):
    #print('date', date)
    date = str(date).strip()
    
    if date:
        
        try:
            return datetime.strptime(date, "%d-%b-%y").strftime('%Y-%m-%d')
        except:
            try: 
                return datetime.strptime(date, "%d-%b-%Y").strftime('%Y-%m-%d')
            except: 
                try:
                    return datetime.strptime(date, "%m/%d/%Y").strftime('%Y-%m-%d')
                except: 
                    try:
                        return datetime.strptime(date, "%m-%d-%Y").strftime('%Y-%m-%d')
                    except:
                        try:
                            return datetime.strptime(date, "%Y-%m-%d %H:%M:%S").strftime('%Y-%m-%d')
                        except:
                            try:
                                return datetime.strptime('', "%m/%d/%Y").strftime('%Y-%m-%d')
                            except:
                                return ''
                        # '2019-07-18 00:00:00' does not match format '%d-%b-%Y'
                        # 2021-01-31 00:00:00
    else:
        return date


def truncate_full_db():
    cursor.execute('''
                EXEC sp_MSforeachtable 'TRUNCATE TABLE ?' 
                ''')
    cursor.commit()
    # quit()


def process_report(file):

    df = pd.read_excel(file)

    df = df[(df['Maxout Date Applicability and Note'].str.contains('prior employer'))|
            (df['Maxout Date Applicability and Note'].str.contains('Prior Employer'))&
            (df['Priority Date']!='')]


    month_list = []
    
    for column in (df.columns):
        if '00:00:00' in str(column):
            month_list.append(column)
            
    # print(month_list)
    # quit()
    for index, row in df.iterrows(): 
        for month in month_list:
                    
            Priority_Date=''
            if 'Priority Date' in df.columns:
                
                Priority_Date = change_format(str(row["Priority Date"]))
            
            Priority_Country=''
            if 'Priority Country' in df.columns:
                Priority_Country = (row["Priority Country"])
                if Priority_Country == 'China- mainland born':
                    Priority_Country = 'China'
                elif Priority_Country == 'All Chargeability Areas Except Those Listed' :
                    Priority_Country = 'All Chargeability'

            Priority_Category=''
            if 'Priority Category' in df.columns:
                Priority_Category = (row["Priority Category"])
                if Priority_Category == 'Employment-1st':
                    Priority_Category = 'EB-1'
                elif Priority_Category == 'Employment-2nd':
                    Priority_Category = 'EB-2'
                elif Priority_Category == 'Employment-3rd':
                    Priority_Category = 'EB-3'
                elif Priority_Category == 'Employment-4th':
                    Priority_Category = 'EB-4'
                elif Priority_Category == 'Employment-5th':
                    Priority_Category = 'EB-5'

            # print(Priority_Category,Priority_Country,month)
            # quit()
            result = cursor.execute('''select * from dbo.VisaBulletinData
                            where Priority_Category like '{}' 
                            and  Priority_Country like '{}'
                            and Priority_Processing_Category = 'Final Action' 
                            and Visa_Bulletin_Month_and_Year = '{}'
                            and Priority_Date  !='1900-01-01 00:00:00.000' '''.format(Priority_Category,Priority_Country,month)).fetchall()
            if result:
                if len(result)>1:
                    print('Database shows more than one record' )        
                    df[month][index] = 'More records '
                else:
                    
                    actual_filing_date = change_format(result[0].Priority_Date) 
                    filed_date = Priority_Date

                    actual_filing_date = datetime.strptime(str(actual_filing_date), "%Y-%m-%d").date()
                    filed_date = datetime.strptime(str(filed_date), "%Y-%m-%d").date()

                    final_filing_date = (actual_filing_date-filed_date).days
                    if final_filing_date <=0:
                        df[month][index] = 'no'
                    else:
                        df[month][index] = 'yes'
            else:
                 df[month][index] = 'No Records'   
    

    date_columns = ['Priority Date','Max Stay'] 
        
    for d_h in date_columns:
        if d_h in df:
            if "1900-01-01" in df[d_h]:
                df[d_h] = ""
            else:
                df[d_h] = pd.to_datetime(df[d_h], format='%Y-%m-%d', errors='coerce').dt.date

    for idx,column in enumerate(df.columns):
    
        if '00:00:00' in str(column):
            month_idx = idx
            break

    
    for index, row in df.iterrows(): 
        x = row[month_idx:]
    #     print(x)
    # quit()
        s= ''.join([i[0] for i in x ])
        # print(s)
        result = re.search(r"y{12}", s)
        if result:
            df['Has the PD been current for 12 Consecutive Months?'][index] = 'yes'
        else:
            df['Has the PD been current for 12 Consecutive Months?'][index] = 'no'
        

    renamed_headers = []


    for i in df.columns:
        if '00:00:00' in str(i):
            x = datetime.strptime(str(i), "%Y-%m-%d %H:%M:%S").strftime('%b-%y')
            i = x
        renamed_headers.append(i)
    # print(renamed_headers)
    # quit()
    
    df.columns = renamed_headers

    output_file_path = 'processed_file\output.xlsx'

    writer = pd.ExcelWriter(output_file_path, engine='xlsxwriter', date_format='m/d/yyyy')
    df.to_excel(writer, "yes_or_no", startrow=0, index=False)
    writer.save()

    ###############################
    # formatting output file

    book = load_workbook(output_file_path)
    writer = pd.ExcelWriter(output_file_path, engine = 'openpyxl')
    writer.book = book

    for x in range(1):
        ws = book[book.sheetnames[x]]
        if ws:
            rows = ws.max_row 
            cols= ws.max_column 

            if x == 0:   
                ws.freeze_panes = ws['D2']
            else:
                ws.freeze_panes = ws['F2']

            for y in range(rows):
                for z in range(cols):

                    ws.cell(row=y+1, column=z+1).font = Font(name = 'Calibri (Body)', size = 11)

                    ws.cell(row=y+1, column=z+1).alignment=Alignment(wrap_text=True, horizontal="center", vertical="justify")

                    ws.cell(row=y+1, column=z+1).font= Font(name = 'Calibri (Body)', size= 11)

                    ws.cell(row=y+1, column=z+1).border= Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                    if y == 0:

                        ws.cell(row=y+1, column=z+1).font = Font(name = 'Calibri',size = 12, color = 'FFFFFF')

        for cl in range(cols):
            if cl <= cols:
                ws.column_dimensions[get_column_letter(cl+1)].width = 15

        for rw in range(rows+1):
            if rw <= rows:
                ws.row_dimensions[rw].height = 30

        table = Table(displayName="Table{}".format(x+1), ref="A1:" + get_column_letter(cols) + str(rows))

        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)

        # for z in range(cols):
        #    ws.cell(row=1, column=z+1).font = Font(size = 12,color = 'ffffff')

    writer.save()
    writer.close()


    

def import_data_to_db(file):

    file_sheets = ['Visa Bulletin Formatted_2011', 'Visa Bulletin Formatted_2012', 'Visa Bulletin Formatted_2013', 'Visa Bulletin Formatted_2014', 'Visa Bulletin Formatted_2015', 'Visa Bulletin Formatted_2016', 'Visa Bulletin Formatted_2017', 'Visa Bulletin Formatted_2018', 'Visa Bulletin Formatted_2019', 'Visa Bulletin Formatted_2020', 'Visa Bulletin Formatted_2021']
    
    for sheet in file_sheets:
        df = pd.read_excel(file,sheet_name=sheet)
        
        # column_list = list(df.columns) not required

        # you need to use index in for loop to access by name else it wont work
        for index, row in df.iterrows(): 
            Visa_Bulletin_Month_and_Year=''
            if 'Visa Bulletin Month and Year' in df.columns:
                Visa_Bulletin_Month_and_Year = row['Visa Bulletin Month and Year']

            Priority_Category=''
            if 'Priority Category' in df.columns:
                Priority_Category = row['Priority Category']

            Priority_Country=''
            if 'Priority Country' in df.columns:
                Priority_Country = row['Priority Country']

            Priority_Type=''
            if 'Priority Type' in df.columns:
                Priority_Type = row['Priority Type']

            Priority_Date=''
            if 'Priority Date' in df.columns:
                
                Priority_Date = change_format(row['Priority Date'])
                # print(Priority_Date)
                # quit()
            Priority_Processing_Category=''
            if 'Priority Processing Category' in df.columns:
                Priority_Processing_Category = row['Priority Processing Category']
            
            USCIS_Filing_Cut_off=''
            if 'USCIS - Filing Cut-off ' in df.columns:
                USCIS_Filing_Cut_off  = row['USCIS - Filing Cut-off ']
           
            if True:
                cursor.execute('''
                insert into VisaBulletinData (Visa_Bulletin_Month_and_Year,Priority_Category,Priority_Country,Priority_Type,Priority_Date,Priority_Processing_Category,USCIS_Filing_Cut_off)
                values ('{}','{}','{}','{}','{}','{}','{}')
                '''.format(Visa_Bulletin_Month_and_Year,Priority_Category,Priority_Country,Priority_Type,Priority_Date,Priority_Processing_Category,USCIS_Filing_Cut_off))

                cursor.commit()
            else:
                pass

def changes_to_db_tables():

    tables = ['[dbo].[VisaBulletinData]']
    for  table in tables:
        table_columns = cursor.execute('''
        SELECT name FROM sys.columns WHERE object_id = OBJECT_ID('{}')
        '''.format(table)).fetchall()
        # the above query gives the list of columns as a queryset

        for i in range(len(table_columns)):     
            cursor.execute('''UPDATE {}
                SET {} = NULL WHERE {} = 'nan' '''.format(table,table_columns[i].name,table_columns[i].name))
            cursor.commit()
            # print(table,table_columns[i].name)
                  
            # cursor.execute('''UPDATE {}
            # SET {} = REPLACE({}, '.0','')'''.format(table,table_columns[i].name,table_columns[i].name))
            # cursor.commit()
        

            
def export_to_excel():
    query = '''select * from VisaBulletinData'''    
    df = pd.read_sql(query,conn)
    for d_h in list(df.columns):
        # if d_h =='Visa_Bulletin_Month_and_Year' or d_h =='Priority_Date':
        if 'Year' in d_h or 'Date' in d_h:
            if "1900-01-01" in df[d_h]:
                        df[d_h] = ""
            else:
                df[d_h] = pd.to_datetime(df[d_h], format='%Y-%m-%d', errors='coerce').dt.date
    writer = pd.ExcelWriter('output.xlsx', engine='xlsxwriter',date_format='m/d/yyyy')
    df.to_excel(writer,'first',index=False)
    writer.save()

def start():
    truncate_full_db()
    # quit()

    for file in glob.glob('source_file/*Chart*'):
        
        import_data_to_db(file)

    for file in glob.glob('source_file/*Status*'):
        
        process_report(file)

print('\nProgram Execution Started..\nIn Progress..\n\n')
start()
# changes_to_db_tables()
# export_to_excel()
print('\nCode Execution Completed - All data imported into DB')

