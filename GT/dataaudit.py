from telnetlib import ENCRYPT
import pyodbc 
import os
from datetime import datetime, date
from datetime import timedelta
from dateutil.relativedelta import relativedelta
import pandas as pd
import chardet
from xlsxwriter import Workbook
import glob

from openpyxl import formatting, styles, Workbook as openpyxl_workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill, colors
from openpyxl.styles.colors import Color, ColorDescriptor
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.styles.fills import Fill
from openpyxl.formatting.rule import CellIsRule, Rule
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


import email, smtplib, ssl

from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time
from cryptography.fernet import Fernet

cwd = os.path.dirname(os.path.realpath(__file__))
os.chdir(cwd)

ENVIRONMENT = 'PRODUCTION' #DEVELOPMENT/PRODUCTION
SEND_MAIL = 'YES' #YES/NO
DB_ENCRYPTION = 'NO' #YES/NO

fernet_key = b'zJD8OVkFNpd5N4fJw6pqaWiDrvybkselSQ0fF9SwXfw='
fernet = Fernet(fernet_key)

#'Server=localhost\SQLEXPRESS;'
if ENVIRONMENT == 'PRODUCTION':
    conn = pyodbc.connect('Driver={ODBC Driver 17 for SQL Server};'
                      'Server=localhost\SQLEXPRESS;'
                      'Database=ReportsAutomation;'
                      'Trusted_Connection=yes;')
    driver_path = os.path.join('C:\ImmiLytics\Automation_UiPath', "chromedriver.exe")
else:
    conn = pyodbc.connect('Driver={ODBC Driver 17 for SQL Server};'
                      'Server=localhost;'
                      'Database=ReportsAutomation2;'
                      'Trusted_Connection=yes;')
    driver_path = ("chromedriver.exe")
cursor = conn.cursor()


def truncate_all():
    cursor.execute("Delete from [Case];\
    Delete from [Beneficiary];\
    truncate table [BeneficiaryPriorityDate];\
    truncate table [BeneficiaryEmployment];\
    truncate table [Petitioner];\
    truncate table [Organization];")
    cursor.commit()

def change_format(date):
    #print('date', date)
    date = str(date).strip()
    
    if date:
        
        try:
            return datetime.strptime(date, "%d-%b-%y").strftime('%Y-%m-%d')
        except:
            try: 
                return datetime.strptime(date, "%d-%b-%Y").strftime('%Y-%m-%d')
            except: 
                try:
                    return datetime.strptime(date, "%m/%d/%Y").strftime('%Y-%m-%d')
                except: 
                    
                    return datetime.strptime('', "%m/%d/%Y").strftime('%Y-%m-%d')
    else:
        return date

def change_display_format(date):
    date = str(date).strip()
    if date:
        try:
            return datetime.strptime(date, "%Y-%m-%d").strftime('%Y-%m-%d') #'%d-%b-%y'
        except:
            return date
    

def start():
    truncate_all()
    current_time = datetime.now() 
    month = str(current_time.month).rjust(2, '0')
    day = str(current_time.day).rjust(2, '0')
    todate = month+''+day+''+str(current_time.year)
    from_name = ''
    for name in glob.glob('Source Data/*Beneficiary*'):
        ##print(os.path.basename(name))
        filename_e = os.path.basename(name)
        filename = os.path.splitext(filename_e)[0]
        extension = os.path.splitext(filename_e)[1]
        ##print(extension)
        if extension == '.csv' and  todate in filename:
            from_name = (filename.split('Data_'))[1].split('_'+str(todate))[0].strip()
            benificiary_file_name = 'Reports Automation_Beneficiary Data_'+str(from_name)+'_'+todate+'.csv'
            if os.path.exists('Source Data/'+benificiary_file_name) and filename_e == benificiary_file_name:
                print('Processing - '+benificiary_file_name)
                process_beneficiary_file('Source Data/'+benificiary_file_name,from_name)

    
    for name in glob.glob('Source Data/*Case*'):
        ##print(os.path.basename(name))
        filename_e = os.path.basename(name)
        filename = os.path.splitext(filename_e)[0]
        extension = os.path.splitext(filename_e)[1]
        ##print(extension)
        if extension == '.csv' and  todate in filename:
            from_name = (filename.split('Data_'))[1].split('_'+str(todate))[0].strip()
            case_file_name = 'Reports Automation_Case Data_'+str(from_name)+'_'+todate+'.csv'

            benificiary_file_name = 'Reports Automation_Beneficiary Data_'+str(from_name)+'_'+todate+'.csv'
            if not os.path.exists('Source Data/'+benificiary_file_name):
                Beneficiary_exists = False
            else:
                Beneficiary_exists = True
            if os.path.exists('Source Data/'+case_file_name)  and filename_e == case_file_name:
                print('Processing - '+case_file_name)
                process_case_file('Source Data/'+case_file_name, from_name, Beneficiary_exists)

    #if from_name:
    #    print('Generating Report - '+from_name)
    #    generate_case_report()
        
        
    
        
def process_beneficiary_file(file_path, from_name):
    with open(file_path,'rb') as f:
        rawdata = b''.join([f.readline() for _ in range(20)])
    enc= chardet.detect(rawdata)['encoding'] #UTF-16

    df = pd.read_csv(file_path, encoding=enc,delimiter='\t')
    list_h = df.columns.tolist()
    total_rows = len(df)
    for index, row in df.iterrows():
        #print(index)
        #print(row['Beneficiary Xref'], row['Organization Xref'])
        #if(index==3):
        #    break
        #    return False

        WorkEmail = ''
        if "Beneficiary Work Email Address" in list_h and not pd.isna(row["Beneficiary Work Email Address"]):
                WorkEmail = str(row["Beneficiary Work Email Address"]).strip()

                
        organization_xref = ''
        if 'Organization Xref' in list_h:
            organization_xref = str(row['Organization Xref']).strip()
            #organization_xref = fernet.encrypt(organization_xref.encode())
        
        organization_name = ''
        if "Organization Name" in list_h:
            organization_name = str(str(row['Organization Name']).replace("'", "")).strip()
            if DB_ENCRYPTION == "YES":
                organization_name = (fernet.encrypt(organization_name.encode())).decode("utf-8") 
                
        organization_id = ''
        if organization_xref and organization_name :
            #print("INSERT INTO dbo.Organization(OrganizationXref, OrganizationName) VALUES ('{}', '{}')".format(organization_xref, organization_name))
            results = cursor.execute("SELECT * FROM dbo.Organization where OrganizationXref='{}'".format(organization_xref)).fetchall()
            length = len(results)
            if length <= 0:
                cursor.execute("INSERT INTO dbo.Organization(OrganizationXref, OrganizationName) VALUES ('{}', '{}')".format(organization_xref, organization_name))
                cursor.execute("SELECT @@IDENTITY AS ID;")
                organization_id = cursor.fetchone()[0]
                cursor.commit()
                ##print('inserted')
            else:
                organization_id = results[0].OrganizationId

        
        petitioner_xref = ''
        if "Petitioner Xref" in list_h:
            petitioner_xref = str(row['Petitioner Xref']).strip()
        
        petitioner_name = ''
        if "Petitioner Name" in list_h:
            petitioner_name = str(str(row['Petitioner Name']).replace("'", "")).strip()
            

        petitioner_id = ''
        is_primary_beneficiary = 1
        if petitioner_xref  and petitioner_name :
            if petitioner_name == 'Individual Client' :
                if str(row['Primary Beneficiary Id']).strip():
                    ##print("SELECT PetitionerId FROM dbo.Beneficiary where BeneficiaryXref='{}'".format(row['Primary Beneficiary Id'].strip()))
                    results = cursor.execute("SELECT PetitionerId FROM dbo.Beneficiary where BeneficiaryXref='{}'".format(str(row['Primary Beneficiary Id']).strip())).fetchall()
                    length = len(results)
                    if length > 0:
                        petitioner_id = results[0][0]
                    is_primary_beneficiary = 0
                        
            else:
                results = cursor.execute("SELECT * FROM dbo.Petitioner where PetitionerXref='{}' and OrganizationId={}".format(petitioner_xref, organization_id)).fetchall()
                length = len(results)
                if length <= 0:
                    ##print("INSERT INTO dbo.Petitioner(PetitionerXref, PetitionerName, OrganizationId) VALUES ('{}', '{}', '{}')")
                    cursor.execute("INSERT INTO dbo.Petitioner(PetitionerXref, PetitionerName, OrganizationId) VALUES ('{}', '{}', '{}')".format(petitioner_xref, petitioner_name, organization_id))
                    cursor.execute("SELECT @@IDENTITY AS ID;")
                    petitioner_id = cursor.fetchone()[0]
                    cursor.commit()
                else:
                    petitioner_id = results[0].PetitionerId
        

        if petitioner_id :
            beneficiary_xref = ''
            if "Beneficiary Xref" in list_h and not pd.isna(row["Beneficiary Xref"]):
                beneficiary_xref = str(row["Beneficiary Xref"]).strip()
            
            beneficiary_type = ''
            if "Beneficiary Type" in list_h and not pd.isna(row["Beneficiary Type"]):
                beneficiary_type = str(row["Beneficiary Type"]).strip()
            
            beneficiary_record_creation_date = ''
            if "Beneficiary Record Creation Date" in list_h and str(row["Beneficiary Record Creation Date"]).strip() and not pd.isna(row["Beneficiary Record Creation Date"]):
                beneficiary_record_creation_date = change_format(row["Beneficiary Record Creation Date"])
            
            beneficiary_record_inactivation_date = ''
            if "Beneficiary Record Inactivation Date" in list_h and str(row["Beneficiary Record Inactivation Date"]).strip() and not pd.isna(row["Beneficiary Record Inactivation Date"]):
                beneficiary_record_inactivation_date = change_format(row["Beneficiary Record Inactivation Date"])

            beneficiary_record_status = 0
            if "Beneficiary Record Status" in list_h and not pd.isna(row["Beneficiary Record Status"]):
                beneficiary_record_status = str(row["Beneficiary Record Status"]).strip()
                if beneficiary_record_status == 'Active':
                    beneficiary_record_status = 1

            beneficiary_employee_id = ''
            if "Beneficiary Employee Id" in list_h and not pd.isna(row["Beneficiary Employee Id"]):
                beneficiary_employee_id = str(row["Beneficiary Employee Id"]).strip()
                
            
            beneficiary_last_name = ''
            if "Beneficiary Last Name" in list_h and not pd.isna(row["Beneficiary Last Name"]):
                beneficiary_last_name = str(str(row["Beneficiary Last Name"]).strip()).replace("'", "")

            beneficiary_first_name = ''
            if "Beneficiary First Name" in list_h  and not pd.isna(row["Beneficiary First Name"]):
                beneficiary_first_name = str(str(row["Beneficiary First Name"]).strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    beneficiary_first_name = (fernet.encrypt(beneficiary_first_name.encode())).decode("utf-8")     

            beneficiary_middle_name = ''
            if "Beneficiary Middle Name" in list_h and not pd.isna(row["Beneficiary Middle Name"]):
                beneficiary_middle_name = str(str(row["Beneficiary Middle Name"]).strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    beneficiary_middle_name = (fernet.encrypt(beneficiary_middle_name.encode())).decode("utf-8")

            primary_beneficiary_id = ''
            if "Primary Beneficiary Id" in list_h and not pd.isna(row["Primary Beneficiary Id"]):
                primary_beneficiary_id = str(row["Primary Beneficiary Id"]).strip()

            #print(primary_beneficiary_id)
            if primary_beneficiary_id == beneficiary_xref:
                is_primary_beneficiary = 1
            else:
                is_primary_beneficiary = 0

            primary_beneficiary_last_name = ''
            if "Primary Beneficiary Last Name" in list_h and not pd.isna(row["Primary Beneficiary Last Name"]):
                primary_beneficiary_last_name = str(str(row["Primary Beneficiary Last Name"]).strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    primary_beneficiary_last_name = (fernet.encrypt(primary_beneficiary_last_name.encode())).decode("utf-8")
            
            primary_beneficiary_first_name = ''
            if "Primary Beneficiary First Name" in list_h and not pd.isna(row["Primary Beneficiary First Name"]):
                primary_beneficiary_first_name = str(str(row["Primary Beneficiary First Name"]).strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    primary_beneficiary_first_name = (fernet.encrypt(primary_beneficiary_first_name.encode())).decode("utf-8")
            
            relation = ''
            if "Relation" in list_h and not pd.isna(row["Relation"]):
                relation = str(str(row["Relation"]).strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    relation = (fernet.encrypt(relation.encode())).decode("utf-8")


            gender = ''
            if "Gender" in list_h and not pd.isna(row["Gender"]):
                gender = str(str(row["Gender"]).strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    gender = (fernet.encrypt(gender.encode())).decode("utf-8")

            date_of_birth = ''
            if "Date of Birth" in list_h and row["Date of Birth"] and not pd.isna(row["Date of Birth"]):
                date_of_birth = change_format(row["Date of Birth"])

            country_of_birth = ''
            if "Country of Birth" in list_h and not pd.isna(row["Country of Birth"]):
                country_of_birth = str(str(row["Country of Birth"]).strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    country_of_birth = (fernet.encrypt(country_of_birth.encode())).decode("utf-8")

            
            country_of_citizenship = ''
            if "Country of Citizenship" in list_h and not pd.isna(row["Country of Citizenship"]):
                country_of_citizenship = row["Country of Citizenship"].replace("'", "")
                if DB_ENCRYPTION == "YES":
                    country_of_citizenship = (fernet.encrypt(country_of_citizenship.encode())).decode("utf-8")

            alien_number = ''
            if "Alien Number" in list_h and not pd.isna(row["Alien Number"]):
                alien_number = str(row["Alien Number"]).strip()
                if DB_ENCRYPTION == "YES":
                    alien_number = (fernet.encrypt(str(alien_number).encode())).decode("utf-8")
            
            date_of_last_entry_into_the_us = ''
            if "Date of Last Entry into the US" in list_h and str(row["Date of Last Entry into the US"]).strip() and not pd.isna(row["Date of Last Entry into the US"]):
                date_of_last_entry_into_the_us = change_format(row["Date of Last Entry into the US"])

            i94_number = ''
            if "I-94 Number" in list_h and not pd.isna(row["I-94 Number"]):
                i94_number = str(row["I-94 Number"]).strip()
                if DB_ENCRYPTION == "YES":
                    i94_number = (fernet.encrypt(i94_number.encode())).decode("utf-8")

            immigration_status = ''
            if "Immigration Status" in list_h and not pd.isna(row["Immigration Status"]):
                immigration_status = row["Immigration Status"].replace("'", "")

            immigration_status_valid_from = ''
            if "Immigration Status Valid From" in list_h and str(row["Immigration Status Valid From"]).strip() and not pd.isna(row["Immigration Status Valid From"]):
                immigration_status_valid_from = change_format(row["Immigration Status Valid From"])


            immigration_status_expiration_status = ''
            if "Immigration Status Expiration Date" in list_h and str(row["Immigration Status Expiration Date"]).strip() and not pd.isna(row["Immigration Status Expiration Date"]):
                if str(row["Immigration Status Expiration Date"]).strip() == 'D/S':
                    immigration_status_expiration_status = 'D/S'
                else:
                    if 'D/S' in str(row["Immigration Status Expiration Date"]).strip():
                        split1 = (str(row["Immigration Status Expiration Date"]).strip()).split('(D/S)')
                        ##print(split1)
                        immigration_status_expiration_status = change_format(split1[0])
                        immigration_status_expiration_status = str(immigration_status_expiration_status)+' (D/S)'
                    else:
                        immigration_status_expiration_status = change_format(str(row["Immigration Status Expiration Date"]).strip())

            i797_approved_date = ''
            if "I-797 Approved Date" in list_h and str(row["I-797 Approved Date"]).strip() and not pd.isna(row["I-797 Approved Date"]):
                i797_approved_date = change_format(row["I-797 Approved Date"])

            i797_status = ''
            if "I-797 Status" in list_h and not pd.isna(row["I-797 Status"]):
                i797_status = str(row["I-797 Status"]).strip()
                if DB_ENCRYPTION == "YES":
                    i797_status = (fernet.encrypt(i797_status.encode())).decode("utf-8")
            
            i797_valid_from = ''
            if "I-797 Valid From" in list_h and str(row["I-797 Valid From"]).strip() and not pd.isna(row["I-797 Valid From"]):
                i797_valid_from = change_format(str(row["I-797 Valid From"]))

            i797_expiration_date = ''
            if "I-797 Expiration Date" in list_h and str(row["I-797 Expiration Date"]).strip() and not pd.isna(row["I-797 Expiration Date"]):
                i797_expiration_date = change_format(row["I-797 Expiration Date"])

            final_niv_status_valid_from = ''
            if "Final NIV-H/L Status Valid From" in list_h and str(row["Final NIV-H/L Status Valid From"]).strip() and not pd.isna(row["Final NIV-H/L Status Valid From"]):
                final_niv_status_valid_from = change_format(row["Final NIV-H/L Status Valid From"])
            
            final_niv_maxout_date = ''
            if "Final NIV (Maxout) Date" in list_h and str(row["Final NIV (Maxout) Date"]).strip() and not pd.isna(row["Final NIV (Maxout) Date"]):
                final_niv_maxout_date = change_format(row["Final NIV (Maxout) Date"])

            maxout_note = ''
            if "Maxout Date Applicability and Note" in list_h and not pd.isna(row["Maxout Date Applicability and Note"]):
                maxout_note = str(str(row["Maxout Date Applicability and Note"]).strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    maxout_note = (fernet.encrypt(maxout_note.encode())).decode("utf-8")

            ped = ''
            if "PED" in list_h and str(row["PED"]).strip() and not pd.isna(row["PED"]):
                ped = change_format(row["PED"])

            ead_type = ''
            if "EAD Type" in list_h and not pd.isna(row["EAD Type"]):
                ead_type = str(str(row["EAD Type"]).strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    ead_type = (fernet.encrypt(ead_type.encode())).decode("utf-8")

            ead_valid_from = ''
            if "EAD Valid From" in list_h and str(row["EAD Valid From"]).strip() and not pd.isna(row["EAD Valid From"]):
                ead_valid_from = change_format(row["EAD Valid From"])

            ead_expiration_date = ''
            if "EAD Expiration Date" in list_h and str(row["EAD Expiration Date"]).strip() and not pd.isna(row["EAD Expiration Date"]):
                ead_expiration_date = change_format(row["EAD Expiration Date"])
            
            ap_valid_from = ''
            if "AP Valid From" in list_h and row["AP Valid From"].strip() and not pd.isna(row["AP Valid From"]):
                ap_valid_from = change_format(row["AP Valid From"])

            ap_expiration_date = ''
            if "AP Expiration Date" in list_h and str(row["AP Expiration Date"]).strip() and not pd.isna(row["AP Expiration Date"]):
                ap_expiration_date = change_format(row["AP Expiration Date"])
            

            ead_ap_type = ''
            if "EAD/AP Type" in list_h and not pd.isna(row["EAD/AP Type"]):
                ead_ap_type = str(row["EAD/AP Type"]).strip()
                if DB_ENCRYPTION == "YES":
                    ead_ap_type = (fernet.encrypt(ead_ap_type.encode())).decode("utf-8")
            
            ead_ap_valid_from = ''
            if "EAD/AP Valid From" in list_h and str(row["EAD/AP Valid From"]).strip() and not pd.isna(row["EAD/AP Valid From"]):
                ead_ap_valid_from = change_format(row["EAD/AP Valid From"])

            ead_ap_expiration_date = ''
            if "EAD/AP Expiration Date" in list_h and str(row["EAD/AP Expiration Date"]).strip() and not pd.isna(row["EAD/AP Expiration Date"]):
                ead_ap_expiration_date = change_format(row["EAD/AP Expiration Date"])

            ds_2019_valid_from = ''
            if "DS-2019 Valid From" in list_h and str(row["DS-2019 Valid From"].strip()) and not pd.isna(row["DS-2019 Valid From"]):
                ds_2019_valid_from = change_format(row["DS-2019 Valid From"])

            ds_2019_expiration_date = ''
            if "DS-2019 Expiration Date" in list_h and str(row["DS-2019 Expiration Date"]).strip() and not pd.isna(row["DS-2019 Expiration Date"]):
                ds_2019_expiration_date = change_format(row["DS-2019 Expiration Date"])

            reentry_permit_expiration_date = ''
            if "Re-Entry Permit Expiration Date" in list_h and row["Re-Entry Permit Expiration Date"] and not pd.isna(row["Re-Entry Permit Expiration Date"]):
                reentry_permit_expiration_date = change_format(row["Re-Entry Permit Expiration Date"])

            green_card_valid_from = ''
            if "Green Card Valid From" in list_h and row["Green Card Valid From"] and not pd.isna(row["Green Card Valid From"]):
                green_card_valid_from = change_format(row["Green Card Valid From"])

            green_card_expiration_date = ''
            if "Green Card Expiration Date" in list_h and row["Green Card Expiration Date"] and not pd.isna(row["Green Card Expiration Date"]):
                green_card_expiration_date = change_format(row["Green Card Expiration Date"])

            passport_last_name = ''
            if "Passport Last Name" in list_h and not pd.isna(row["Passport Last Name"]):
                passport_last_name = str(str(row["Passport Last Name"]).strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    passport_last_name = (fernet.encrypt(passport_last_name.encode())).decode("utf-8")

            passport_first_name = ''
            if "Passport First Name" in list_h and not pd.isna(row["Passport First Name"]):
                passport_first_name = str(str(row["Passport First Name"]).strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    passport_first_name = (fernet.encrypt(passport_first_name.encode())).decode("utf-8")
            
            passport_middle_name = ''
            if "Passport Middle Name" in list_h and not pd.isna(row["Passport Middle Name"]):
                passport_middle_name = str(str(row["Passport Middle Name"]).strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    passport_middle_name = (fernet.encrypt(passport_middle_name.encode())).decode("utf-8")

            passport_number = ''
            if "Passport Number" in list_h and not pd.isna(row["Passport Number"]):
                passport_number = str(row["Passport Number"]).strip()
                if DB_ENCRYPTION == "YES":
                    passport_number = (fernet.encrypt(passport_number.encode())).decode("utf-8")

            passport_issuing_country = ''
            if "Passport Issuing Country" in list_h and not pd.isna(row["Passport Issuing Country"]):
                passport_issuing_country = str(str(row["Passport Issuing Country"]).strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    passport_issuing_country = (fernet.encrypt(passport_issuing_country.encode())).decode("utf-8")

            passport_valid_from = ''
            if "Passport Valid From" in list_h and row["Passport Valid From"] and not pd.isna(row["Passport Valid From"]):
                passport_valid_from = change_format(row["Passport Valid From"])
                if DB_ENCRYPTION == "YES":
                    ead_ap_type = (fernet.encrypt(ead_ap_type.encode())).decode("utf-8")

            passport_expiration_date = ''
            if "Passport Expiration Date" in list_h and row["Passport Expiration Date"] and not pd.isna(row["Passport Expiration Date"]):
                passport_expiration_date = change_format(row["Passport Expiration Date"])


            visa_type = ''
            if "Visa Type" in list_h and not pd.isna(row["Visa Type"]):
                visa_type = str(row["Visa Type"]).strip()
                if DB_ENCRYPTION == "YES":
                    visa_type = (fernet.encrypt(visa_type.encode())).decode("utf-8")

            visa_valid_from = ''
            if "Visa Valid From" in list_h and row["Visa Valid From"] and not pd.isna(row["Visa Valid From"]):
                visa_valid_from = change_format(row["Visa Valid From"])

            visa_expiration_date = ''
            if "Visa Expiration Date" in list_h and row["Visa Expiration Date"] and not pd.isna(row["Visa Expiration Date"]):
                visa_expiration_date = change_format(row["Visa Expiration Date"])
            
            employee_hire_date = ''
            if "Employee Hire Date" in list_h and row["Employee Hire Date"] and not pd.isna(row["Employee Hire Date"]):
                employee_hire_date = change_format(row["Employee Hire Date"])
            
            current_job_title = ''
            if "Current Job Title" in list_h and not pd.isna(row["Current Job Title"]):
                current_job_title = str(str(row["Current Job Title"]).strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    current_job_title = (fernet.encrypt(current_job_title.encode())).decode("utf-8")

            work_address_street = ''
            if "Work Address-Street" in list_h and not pd.isna(row["Work Address-Street"]):
                work_address_street = str(str(row["Work Address-Street"]).strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    work_address_street = (fernet.encrypt(work_address_street.encode())).decode("utf-8")

            work_address_city = ''
            if "Work Address-City" in list_h and not pd.isna(row["Work Address-City"]):
                work_address_city = str(str(row["Work Address-City"]).strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    work_address_city = (fernet.encrypt(work_address_city.encode())).decode("utf-8")

            work_address_state = ''
            if "Work Address-State" in list_h and not pd.isna(row["Work Address-State"]):
                work_address_state = str(str(row["Work Address-State"]).strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    work_address_state = (fernet.encrypt(work_address_state.encode())).decode("utf-8")

            work_address_zip = ''
            if "Work Address-Zip" in list_h and not pd.isna(row["Work Address-Zip"]):
                work_address_zip = str(str(row["Work Address-Zip"]).strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    work_address_zip = (fernet.encrypt(work_address_zip.encode())).decode("utf-8")
            
            work_address_country = ''
            if "Work Address-Country" in list_h and not pd.isna(row["Work Address-Country"]):
                work_address_country = str(row["Work Address-Country"].strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    work_address_country = (fernet.encrypt(work_address_country.encode())).decode("utf-8")

            priority_date_1_date = ''
            if "Priority Date 1-Date" in list_h and row["Priority Date 1-Date"] and not pd.isna(row["Priority Date 1-Date"]):
                priority_date_1_date = change_format(row["Priority Date 1-Date"])
                
            
            priority_date_1_category = ''
            if "Priority Date 1-Category" in list_h and not pd.isna(row["Priority Date 1-Category"]):
                priority_date_1_category = str(str(row["Priority Date 1-Category"]).strip()).replace("'", "")

            priority_date_1_country_of_charge = ''
            if "Priority Date 1-Country of Chargeability" in list_h and not pd.isna(row["Priority Date 1-Country of Chargeability"]):
                priority_date_1_country_of_charge = str(row["Priority Date 1-Country of Chargeability"]).strip()

            priority_date_2_date = ''
            if "Priority Date 2-Date" in list_h and row["Priority Date 2-Date"].strip() and not pd.isna(row["Priority Date 2-Date"]):
                priority_date_2_date = change_format(row["Priority Date 2-Date"]).replace("'", "")
            
            priority_date_2_category = ''
            if "Priority Date 2-Category" in list_h and not pd.isna(row["Priority Date 2-Category"]):
                priority_date_2_category = str(str(row["Priority Date 2-Category"]).strip()).replace("'", "")

            priority_date_2_country_of_charge = ''
            if "Priority Date 2-Country of Chargeability" in list_h and not pd.isna(row["Priority Date 2-Country of Chargeability"]):
                priority_date_2_country_of_charge = str(str(row["Priority Date 2-Country of Chargeability"]).strip()).replace("'", "")

            priority_date_3_date = ''
            if "Priority Date 3-Date" in list_h and row["Priority Date 3-Date"].strip() and not pd.isna(row["Priority Date 3-Date"]):
                priority_date_3_date = change_format(row["Priority Date 3-Date"])
            
            priority_date_3_category = ''
            if "Priority Date 3-Category" in list_h and not pd.isna(row["Priority Date 3-Category"]):
                priority_date_3_category = str(str(row["Priority Date 3-Category"]).strip()).replace("'", "")

            priority_date_3_country_of_charge = ''
            if "Priority Date 3-Country of Chargeability" in list_h and not pd.isna(row["Priority Date 3-Country of Chargeability"]):
                priority_date_3_country_of_charge = str(str(row["Priority Date 3-Country of Chargeability"]).strip()).replace("'", "")

            priority_date_4_date = ''
            if "Priority Date 4-Date" in list_h and row["Priority Date 4-Date"] and not pd.isna(row["Priority Date 4-Date"]):
                priority_date_4_date = change_format(row["Priority Date 4-Date"])
            
            priority_date_4_category = ''
            if "Priority Date 4-Category" in list_h and not pd.isna(row["Priority Date 4-Category"]):
                priority_date_4_category = str(str(row["Priority Date 4-Category"]).strip()).replace("'", "")

            priority_date_4_country_of_charge = ''
            if "Priority Date 4-Country of Chargeability" in list_h and not pd.isna(row["Priority Date 4-Country of Chargeability"]):
                priority_date_4_country_of_charge = str(str(row["Priority Date 4-Country of Chargeability"]).strip()).replace("'", "")

            priority_date_5_date = ''
            if "Priority Date 5-Date" in list_h and row["Priority Date 5-Date"] and not pd.isna(row["Priority Date 5-Date"]):
                priority_date_5_date = change_format(row["Priority Date 5-Date"])
            
            priority_date_5_category = ''
            if "Priority Date 5-Category" in list_h and not pd.isna(row["Priority Date 5-Category"]):
                priority_date_5_category = str(str(row["Priority Date 5-Category"]).strip()).replace("'", "")

            priority_date_5_country_of_charge = ''
            if "Priority Date 5-Country of Chargeability" in list_h and not pd.isna(row["Priority Date 5-Country of Chargeability"]):
                priority_date_5_country_of_charge = str(str(row["Priority Date 5-Country of Chargeability"]).strip()).replace("'", "")

            beneficiary_id = ''
            if beneficiary_xref:
                results = cursor.execute("SELECT * FROM dbo.Beneficiary where BeneficiaryXref='{}' and from_name='{}'".format(beneficiary_xref, from_name)).fetchall()
                length = len(results)
                if length <= 0:
                    cursor.execute("INSERT INTO dbo.Beneficiary(WorkEmail,PetitionerId, BeneficiaryXref, BeneficiaryType, SourceCreatedDate, IsActive, InactiveDate, LastName, FirstName, MiddleName, PrimaryBeneficiaryXref, PrimaryBeneficiaryLastName, PrimaryBeneficiaryFirstName, RelationType, Gender, BirthDate, BirthCountry, CitizenshipCountry, AlienNumber, MostRecentUSEntryDate, I94Number, ImmigrationStatus, ImmigrationStatusValidFromDate, ImmigrationStatusExpirationDate, MostRecentI797IssueApprovalDate, MostRecentI797Status, MostRecentI797ValidFromDate, I797ExpirationDate, InitialHlEntryDate, FinalNivDate, MaxOutDateNote, EadType, VisaPedDate, EadValidFromDate, EadExpirationDate, AdvanceParoleValidFromDate, AdvanceParoleExpirationDate, EADAPType, EadApValidFromDate, EadApExpirationDate, Ds2019ValidFromDate, Ds2019ExpirationDate, ReEntryPermitExpirationDate, GreenCardValidFromDate, GreenCardExpirationDate, MostRecentPassportLastName, MostRecentPassportFirstName, MostRecentPassportMiddleName, MostRecentPassportNumber, MostRecentPassportIssuingCountry, MostRecentPassportValidFromDate, MostRecentPassportExpirationDate, VisaType, VisaValidFromDate, VisaExpirationDate, from_name, is_primary_beneficiary, from_source  ) VALUES ('{}','{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}','{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}','{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}','{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}','{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}')".format(WorkEmail,petitioner_id, beneficiary_xref, beneficiary_type, beneficiary_record_creation_date, beneficiary_record_status, beneficiary_record_inactivation_date, beneficiary_last_name, beneficiary_first_name, beneficiary_middle_name, primary_beneficiary_id, primary_beneficiary_last_name, primary_beneficiary_first_name, relation, gender, date_of_birth, country_of_birth, country_of_citizenship,  alien_number, date_of_last_entry_into_the_us, i94_number, immigration_status, immigration_status_valid_from, immigration_status_expiration_status, i797_approved_date, i797_status, i797_valid_from, i797_expiration_date,  final_niv_status_valid_from, final_niv_maxout_date, maxout_note, ead_type, ped, ead_valid_from, ead_expiration_date, ap_valid_from, ap_expiration_date, ead_ap_type, ead_ap_valid_from, ead_ap_expiration_date, ds_2019_valid_from, ds_2019_expiration_date, reentry_permit_expiration_date, green_card_valid_from, green_card_expiration_date, passport_last_name, passport_first_name, passport_middle_name, passport_number, passport_issuing_country, passport_valid_from, passport_expiration_date, visa_type, visa_valid_from, visa_expiration_date, from_name, is_primary_beneficiary, 'beneficiary'))
                    cursor.execute("SELECT @@IDENTITY AS ID;")
                    beneficiary_id = cursor.fetchone()[0]
                    cursor.commit()
                else:

                    beneficiary_id = results[0].BeneficiaryId
                    
                    cursor.execute("UPDATE dbo.Beneficiary SET WorkEmail='{}',PetitionerId='{}', BeneficiaryXref='{}', BeneficiaryType='{}', SourceCreatedDate='{}', IsActive='{}', InactiveDate='{}', LastName='{}', FirstName='{}', MiddleName='{}', PrimaryBeneficiaryXref='{}', PrimaryBeneficiaryLastName='{}', PrimaryBeneficiaryFirstName='{}', RelationType='{}', Gender='{}', BirthDate='{}', BirthCountry='{}', CitizenshipCountry='{}', AlienNumber='{}', MostRecentUSEntryDate='{}', I94Number='{}', ImmigrationStatus='{}', ImmigrationStatusValidFromDate='{}', ImmigrationStatusExpirationDate='{}', MostRecentI797IssueApprovalDate='{}', MostRecentI797Status='{}', MostRecentI797ValidFromDate='{}', I797ExpirationDate='{}', InitialHlEntryDate='{}', FinalNivDate='{}', MaxOutDateNote='{}', EadType='{}', VisaPedDate='{}', EadValidFromDate='{}', EadExpirationDate='{}', AdvanceParoleValidFromDate='{}', AdvanceParoleExpirationDate='{}', EADAPType='{}', EadApValidFromDate='{}', EadApExpirationDate='{}', Ds2019ValidFromDate='{}', Ds2019ExpirationDate='{}', ReEntryPermitExpirationDate='{}', GreenCardValidFromDate='{}', GreenCardExpirationDate='{}', MostRecentPassportLastName='{}', MostRecentPassportFirstName='{}', MostRecentPassportMiddleName='{}', MostRecentPassportNumber='{}', MostRecentPassportIssuingCountry='{}', MostRecentPassportValidFromDate='{}', MostRecentPassportExpirationDate='{}', VisaType='{}', VisaValidFromDate='{}', VisaExpirationDate='{}', from_name='{}', is_primary_beneficiary='{}' WHERE BeneficiaryId='{}' ".format(WorkEmail,petitioner_id, beneficiary_xref, beneficiary_type, beneficiary_record_creation_date, beneficiary_record_status, beneficiary_record_inactivation_date, beneficiary_last_name, beneficiary_first_name, beneficiary_middle_name, primary_beneficiary_id, primary_beneficiary_last_name, primary_beneficiary_first_name, relation, gender, date_of_birth, country_of_birth, country_of_citizenship,  alien_number, date_of_last_entry_into_the_us, i94_number, immigration_status, immigration_status_valid_from, immigration_status_expiration_status, i797_approved_date, i797_status, i797_valid_from, i797_expiration_date,  final_niv_status_valid_from, final_niv_maxout_date, maxout_note, ead_type, ped, ead_valid_from, ead_expiration_date, ap_valid_from, ap_expiration_date, ead_ap_type, ead_ap_valid_from, ead_ap_expiration_date, ds_2019_valid_from, ds_2019_expiration_date, reentry_permit_expiration_date, green_card_valid_from, green_card_expiration_date, passport_last_name, passport_first_name, passport_middle_name, passport_number, passport_issuing_country, passport_valid_from, passport_expiration_date, visa_type, visa_valid_from, visa_expiration_date, from_name, is_primary_beneficiary, beneficiary_id))
                    cursor.commit()
            
            if beneficiary_id:
                results = cursor.execute("SELECT * FROM dbo.BeneficiaryPriorityDate where BeneficiaryId='{}'".format(beneficiary_id)).fetchall()
                length = len(results)
                if length <= 0:
                    cursor.execute("INSERT INTO dbo.BeneficiaryPriorityDate(BeneficiaryId, Priority1Date, Priority1Category, Priority1Country, Priority2Date, Priority2Category, Priority2Country, Priority3Date, Priority3Category, Priority3Country, Priority4Date, Priority4Category, Priority4Country, Priority5Date, Priority5Category, Priority5Country) VALUES ('{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}')".format(beneficiary_id, priority_date_1_date, priority_date_1_category, priority_date_1_country_of_charge, priority_date_2_date, priority_date_2_category, priority_date_2_country_of_charge, priority_date_3_date, priority_date_3_category, priority_date_3_country_of_charge, priority_date_4_date, priority_date_4_category, priority_date_4_country_of_charge, priority_date_5_date, priority_date_5_category, priority_date_5_country_of_charge))
                    cursor.commit()
                else:
                    cursor.execute("UPDATE dbo.BeneficiaryPriorityDate SET BeneficiaryId='{}', Priority1Date='{}', Priority1Category='{}', Priority1Country='{}', Priority2Date='{}', Priority2Category='{}', Priority2Country='{}', Priority3Date='{}', Priority3Category='{}', Priority3Country='{}', Priority4Date='{}', Priority4Category='{}', Priority4Country='{}', Priority5Date='{}', Priority5Category='{}', Priority5Country='{}' WHERE BeneficiaryId='{}'".format(beneficiary_id, priority_date_1_date, priority_date_1_category, priority_date_1_country_of_charge, priority_date_2_date, priority_date_2_category, priority_date_2_country_of_charge, priority_date_3_date, priority_date_3_category, priority_date_3_country_of_charge, priority_date_4_date, priority_date_4_category, priority_date_4_country_of_charge, priority_date_5_date, priority_date_5_category, priority_date_5_country_of_charge, beneficiary_id))
                    cursor.commit()

            if beneficiary_id:
                results = cursor.execute("SELECT * FROM dbo.BeneficiaryEmployment where BeneficiaryId='{}'".format(beneficiary_id)).fetchall()
                length = len(results)
                if length <= 0:
                    cursor.execute("INSERT INTO dbo.BeneficiaryEmployment(BeneficiaryId, EmployeeId, HireDate, JobTitle, Address1, City, StateProvince, ZipCode, Country) VALUES ('{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}' )".format(beneficiary_id, beneficiary_employee_id, employee_hire_date, current_job_title, work_address_street, work_address_city, work_address_state, work_address_zip, work_address_country))
                    cursor.commit()
                else:
                    cursor.execute("UPDATE dbo.BeneficiaryEmployment SET BeneficiaryId='{}', EmployeeId='{}', HireDate='{}', JobTitle='{}', Address1='{}', City='{}', StateProvince='{}', ZipCode='{}', Country='{}' WHERE BeneficiaryId='{}'".format(beneficiary_id, beneficiary_employee_id, employee_hire_date, current_job_title, work_address_street, work_address_city, work_address_state, work_address_zip, work_address_country, beneficiary_id))
                    cursor.commit()
            

def process_case_file(file_path, from_name, Beneficiary_exists):
    with open(file_path,'rb') as f:
        rawdata = b''.join([f.readline() for _ in range(20)])
    enc= chardet.detect(rawdata)['encoding'] #UTF-16

    df = pd.read_csv(file_path, encoding=enc,delimiter='\t')
    list_h = df.columns.tolist()
    total_rows = len(df)

    if Beneficiary_exists==False:
        from_source = 'case' 
    else:
        from_source = 'beneficiary' 
    for index, row in df.iterrows():
        organization_xref = ''
        if 'Organization Xref' in list_h:
            organization_xref = str(row['Organization Xref']).strip()
        
        organization_name = ''
        if "Organization Name" in list_h:
            organization_name = str(str(row['Organization Name']).replace("'", "")).strip()
            if DB_ENCRYPTION == "YES":
                    organization_name = (fernet.encrypt(organization_name.encode())).decode("utf-8")

        organization_id = ''
        if organization_xref  and organization_name :
            ##print("SELECT * FROM dbo.Organization where OrganizationXref='{}' and OrganizationName = '{}'".format(organization_xref, organization_name))
            results = cursor.execute("SELECT * FROM dbo.Organization where OrganizationXref='{}'".format(organization_xref)).fetchall()
            length = len(results)
            if length <= 0:
                ##print("INSERT INTO dbo.Organization(OrganizationXref, OrganizationName) VALUES ('{}', '{}')".format(organization_xref, organization_name))
                cursor.execute("INSERT INTO dbo.Organization(OrganizationXref, OrganizationName) VALUES ('{}', '{}')".format(organization_xref, organization_name))
                cursor.execute("SELECT @@IDENTITY AS ID;")
                organization_id = cursor.fetchone()[0]
                cursor.commit()
                ##print('inserted')
            else:
                organization_id = results[0].OrganizationId
        
        ##print('oid ', organization_id)
        petitioner_xref = ''
        if "Petitioner Xref" in list_h:
            petitioner_xref = str(row['Petitioner Xref']).strip()
        
        petitioner_name = ''
        if "Petitioner Name" in list_h:
            petitioner_name = str(str(row['Petitioner Name']).replace("'", "")).strip()
            

        petitioner_id = ''
        is_primary_beneficiary = 1
        if petitioner_xref and petitioner_name:
            if petitioner_name == 'Individual Client' :
                if str(row['Primary Beneficiary Xref']).strip():
                    ##print("SELECT PetitionerId FROM dbo.Beneficiary where BeneficiaryXref='{}'".format(row['Primary Beneficiary Xref'].strip()))
                    results = cursor.execute("SELECT PetitionerId FROM dbo.Beneficiary where BeneficiaryXref='{}'".format(str(row['Primary Beneficiary Xref']).strip())).fetchall()
                    length = len(results)
                    if length > 0:
                        petitioner_id = results[0][0]
                        is_primary_beneficiary = 0
                   
                    

            else:
                ##print("SELECT * FROM dbo.Petitioner where PetitionerXref='{}' and PetitionerName = '{}' and OrganizationId={}".format(petitioner_xref, petitioner_name, organization_id))
                results = cursor.execute("SELECT * FROM dbo.Petitioner where PetitionerXref='{}' and OrganizationId={}".format(petitioner_xref,  organization_id)).fetchall()
                length = len(results)
                if length <= 0:
                    ##print("INSERT INTO dbo.Petitioner(PetitionerXref, PetitionerName, OrganizationId) VALUES ('{}', '{}', '{}')".format(petitioner_xref, petitioner_name, organization_id))
                    cursor.execute("INSERT INTO dbo.Petitioner(PetitionerXref, PetitionerName, OrganizationId) VALUES ('{}', '{}', '{}')".format(petitioner_xref, petitioner_name, organization_id))
                    cursor.execute("SELECT @@IDENTITY AS ID;")
                    petitioner_id = cursor.fetchone()[0]
                    cursor.commit()
                else:
                    petitioner_id = results[0].PetitionerId
        
        ##print('pid ', petitioner_id)
        #if petitioner_id :
        if True:
            beneficiary_xref = ''
            if "Beneficiary Xref" in list_h and not pd.isna(row["Beneficiary Xref"]):
                beneficiary_xref = str(row["Beneficiary Xref"]).strip()
            
            beneficiary_type = ''
            if "Beneficiary Type" in list_h and not pd.isna(row["Beneficiary Type"]):
                beneficiary_type = str(row["Beneficiary Type"]).strip()
            
            beneficiary_record_creation_date = ''
            if "Beneficiary Record Creation Date" in list_h and row["Beneficiary Record Creation Date"] and not pd.isna(row["Beneficiary Record Creation Date"]):
                beneficiary_record_creation_date = change_format(row["Beneficiary Record Creation Date"])
            
            beneficiary_record_inactivation_date = ''
            if "Beneficiary Record Inactivation Date" in list_h and row["Beneficiary Record Inactivation Date"] and not pd.isna(row["Beneficiary Record Inactivation Date"]):
                beneficiary_record_inactivation_date = change_format(row["Beneficiary Record Inactivation Date"])

            beneficiary_record_status = 0
            if "Beneficiary Record Status" in list_h and not pd.isna(row["Beneficiary Record Status"]):
                beneficiary_record_status = str(row["Beneficiary Record Status"]).strip()
                if beneficiary_record_status == 'Active':
                    beneficiary_record_status = 1
                else:
                    beneficiary_record_status = 0

            beneficiary_last_name = ''
            if "Beneficiary Last Name" in list_h and not pd.isna(row["Beneficiary Last Name"]):
                beneficiary_last_name = str(str(row["Beneficiary Last Name"]).strip()).replace("'", "")

            beneficiary_first_name = ''
            if "Beneficiary First Name" in list_h  and not pd.isna(row["Beneficiary First Name"]):
                beneficiary_first_name = str(str(row["Beneficiary First Name"]).strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    beneficiary_first_name = (fernet.encrypt(beneficiary_first_name.encode())).decode("utf-8")

            beneficiary_middle_name = ''
            if "Beneficiary Middle Name" in list_h and not pd.isna(row["Beneficiary Middle Name"]):
                beneficiary_middle_name = str(str(row["Beneficiary Middle Name"]).strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    beneficiary_middle_name = (fernet.encrypt(beneficiary_middle_name.encode())).decode("utf-8")

            primary_beneficiary_id = ''
            if "Primary Beneficiary Xref" in list_h and not pd.isna(row["Primary Beneficiary Xref"]):
                primary_beneficiary_id = str(row["Primary Beneficiary Xref"]).strip()

            if primary_beneficiary_id == beneficiary_xref:
                is_primary_beneficiary = 1
            else:
                is_primary_beneficiary = 0

            primary_beneficiary_last_name = ''
            if "Primary Beneficiary Last Name" in list_h and not pd.isna(row["Primary Beneficiary Last Name"]):
                primary_beneficiary_last_name = (row["Primary Beneficiary Last Name"].strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    primary_beneficiary_last_name = (fernet.encrypt(primary_beneficiary_last_name.encode())).decode("utf-8")
            
            primary_beneficiary_first_name = ''
            if "Primary Beneficiary First Name" in list_h and not pd.isna(row["Primary Beneficiary First Name"]):
                primary_beneficiary_first_name = (row["Primary Beneficiary First Name"].strip()).replace("'", "")
                if DB_ENCRYPTION == "YES":
                    primary_beneficiary_first_name = (fernet.encrypt(primary_beneficiary_first_name.encode())).decode("utf-8")
            
            relation = ''
            if "Relation" in list_h and not pd.isna(row["Relation"]):
                relation = row["Relation"].strip()
                if DB_ENCRYPTION == "YES":
                    relation = (fernet.encrypt(relation.encode())).decode("utf-8")

            immigration_status = ''
            if "Immigration Status" in list_h and not pd.isna(row["Immigration Status"]):
                immigration_status = str(row["Immigration Status"]).strip()

            immigration_status_expiration_status = ''
            if "Immigration Status Expiration Date" in list_h and row["Immigration Status Expiration Date"] and not pd.isna(row["Immigration Status Expiration Date"]):
                if row["Immigration Status Expiration Date"].strip() == 'D/S':
                    immigration_status_expiration_status = 'D/S'
                else:
                    if 'D/S' in row["Immigration Status Expiration Date"]:
                        split1 = str(str(row["Immigration Status Expiration Date"]).strip()).split('(D/S)')
                        immigration_status_expiration_status = change_format(split1[0])
                        immigration_status_expiration_status = str(immigration_status_expiration_status)+' (D/S)'
                    else:
                        immigration_status_expiration_status = change_format(row["Immigration Status Expiration Date"])

            i797_approved_date = ''
            if "I-797 Approved Date" in list_h and row["I-797 Approved Date"] and not pd.isna(row["I-797 Approved Date"]):
                i797_approved_date = change_format(row["I-797 Approved Date"])

            i797_status = ''
            if "I-797 Status" in list_h and not pd.isna(row["I-797 Status"]):
                i797_status = str(row["I-797 Status"]).strip()
                if DB_ENCRYPTION == "YES":
                    i797_status = (fernet.encrypt(i797_status.encode())).decode("utf-8")
            
            i797_expiration_date = ''
            if "I-797 Expiration Date" in list_h and row["I-797 Expiration Date"] and not pd.isna(row["I-797 Expiration Date"]):
                i797_expiration_date = change_format(row["I-797 Expiration Date"])

            final_niv_maxout_date = ''
            if "Final NIV (Maxout) Date" in list_h and row["Final NIV (Maxout) Date"] and not pd.isna(row["Final NIV (Maxout) Date"]):
                final_niv_maxout_date = change_format(row["Final NIV (Maxout) Date"])

            maxout_note = ''
            if "Maxout Date Applicability and Note" in list_h and not pd.isna(row["Maxout Date Applicability and Note"]):
                maxout_note = str(row["Maxout Date Applicability and Note"]).strip()
                if DB_ENCRYPTION == "YES":
                    maxout_note = (fernet.encrypt(maxout_note.encode())).decode("utf-8")

            beneficiary_id = ''
            if beneficiary_xref:
                results = cursor.execute("SELECT * FROM dbo.Beneficiary where BeneficiaryXref='{}' and from_name='{}'".format(beneficiary_xref, from_name)).fetchall()
                length = len(results)
                if length <= 0:
                    
                    cursor.execute("INSERT INTO dbo.Beneficiary(PetitionerId, BeneficiaryXref, BeneficiaryType, SourceCreatedDate, IsActive, InactiveDate, LastName, FirstName, MiddleName, PrimaryBeneficiaryXref, PrimaryBeneficiaryLastName, PrimaryBeneficiaryFirstName, RelationType, ImmigrationStatus, ImmigrationStatusExpirationDate, MostRecentI797IssueApprovalDate, MostRecentI797Status, I797ExpirationDate, FinalNivDate, MaxOutDateNote, from_name, is_primary_beneficiary, from_source  ) VALUES ('{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}','{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}')".format(petitioner_id, beneficiary_xref, beneficiary_type, beneficiary_record_creation_date, beneficiary_record_status, beneficiary_record_inactivation_date, beneficiary_last_name, beneficiary_first_name, beneficiary_middle_name, primary_beneficiary_id, primary_beneficiary_last_name, primary_beneficiary_first_name, relation, immigration_status, immigration_status_expiration_status, i797_approved_date, i797_status, i797_expiration_date, final_niv_maxout_date, maxout_note, from_name, is_primary_beneficiary, from_source))
                    cursor.execute("SELECT @@IDENTITY AS ID;")
                    beneficiary_id = cursor.fetchone()[0]
                    cursor.commit()
                else:
                    beneficiary_id = results[0].BeneficiaryId
                    cursor.execute("UPDATE  dbo.Beneficiary SET PetitionerId='{}', BeneficiaryXref='{}', BeneficiaryType='{}', SourceCreatedDate='{}', IsActive='{}', InactiveDate='{}', LastName='{}', FirstName='{}', MiddleName='{}', PrimaryBeneficiaryXref='{}', PrimaryBeneficiaryLastName='{}', PrimaryBeneficiaryFirstName='{}', RelationType='{}', ImmigrationStatus='{}', ImmigrationStatusExpirationDate='{}', MostRecentI797IssueApprovalDate='{}', MostRecentI797Status='{}', I797ExpirationDate='{}', FinalNivDate='{}', MaxOutDateNote='{}', from_name='{}', is_primary_beneficiary='{}' WHERE BeneficiaryId='{}'  ".format(petitioner_id, beneficiary_xref, beneficiary_type, beneficiary_record_creation_date, beneficiary_record_status, beneficiary_record_inactivation_date, beneficiary_last_name, beneficiary_first_name, beneficiary_middle_name, primary_beneficiary_id, primary_beneficiary_last_name, primary_beneficiary_first_name, relation, immigration_status, immigration_status_expiration_status, i797_approved_date, i797_status, i797_expiration_date, final_niv_maxout_date, maxout_note, from_name, is_primary_beneficiary, beneficiary_id))
                    cursor.commit()

            ##print('bid ',beneficiary_id)
            if beneficiary_id:
                case_xref = ''
                if "Case Xref" in list_h and not pd.isna(row["Case Xref"]):
                    case_xref = str(row["Case Xref"]).strip()
                
                case_creation_date = ''
                if "Case Created Date" in list_h and row["Case Created Date"] and not pd.isna(row["Case Created Date"]):
                    case_creation_date = change_format(row["Case Created Date"])

                case_petition_name = ''
                if "Case Petition Name" in list_h and not pd.isna(row["Case Petition Name"]):
                    case_petition_name = str(str(row["Case Petition Name"]).strip()).replace("'", "")
                    if DB_ENCRYPTION == "YES":
                        case_petition_name = (fernet.encrypt(case_petition_name.encode())).decode("utf-8")

                case_type = ''
                if "Case Type" in list_h and not pd.isna(row["Case Type"]):
                    case_type = str(str(row["Case Type"]).strip()).replace("'", "")
                    if DB_ENCRYPTION == "YES":
                        case_type = (fernet.encrypt(case_type.encode())).decode("utf-8")

                case_description = ''
                if "Case Description" in list_h and not pd.isna(row["Case Description"]):
                    case_description = str(str(row["Case Description"]).strip()).replace("'", "")
                    if DB_ENCRYPTION == "YES":
                        case_description = (fernet.encrypt(case_description.encode())).decode("utf-8")
                
                case_filed_date = ''
                if "Case Filed Date" in list_h and row["Case Filed Date"] and not pd.isna(row["Case Filed Date"]):
                    case_filed_date = change_format(row["Case Filed Date"])
                
                
                case_receipt_number = ''
                if "Case Receipt Number" in list_h and not pd.isna(row["Case Receipt Number"]):
                    case_receipt_number = str(row["Case Receipt Number"]).strip()
                    if DB_ENCRYPTION == "YES":
                        case_receipt_number = (fernet.encrypt(case_receipt_number.encode())).decode("utf-8")

                case_receipt_status = ''
                if "Case Receipt Status" in list_h and not pd.isna(row["Case Receipt Status"]):
                    case_receipt_status = str(row["Case Receipt Status"]).strip()
                    if DB_ENCRYPTION == "YES":
                        case_receipt_status = (fernet.encrypt(case_receipt_status.encode())).decode("utf-8")

                rfe_audit_received_date = ''
                if "RFE/Audit Received Date" in list_h and row["RFE/Audit Received Date"] and not pd.isna(row["RFE/Audit Received Date"]):
                    rfe_audit_received_date = change_format(row["RFE/Audit Received Date"])
                
                rfe_audit_due_date = ''
                if "RFE/Audit Response Due Date" in list_h and row["RFE/Audit Response Due Date"] and not pd.isna(row["RFE/Audit Response Due Date"]):
                    rfe_audit_due_date = change_format(row["RFE/Audit Response Due Date"])
                
                rfe_audit_submitted_date = ''
                if "RFE/Audit Response Submitted Date" in list_h and row["RFE/Audit Response Submitted Date"] and not pd.isna(row["RFE/Audit Response Submitted Date"]):
                    rfe_audit_submitted_date = change_format(row["RFE/Audit Response Submitted Date"])

                primary_case_status = ''
                if "Primary Case Status" in list_h and not pd.isna(row["Primary Case Status"]):
                    primary_case_status = str(row["Primary Case Status"]).strip()

                secondary_case_status = ''
                if "Secondary Case Status" in list_h and not pd.isna(row["Secondary Case Status"]):
                    secondary_case_status = str(str(row["Secondary Case Status"]).strip()).replace("'", "")
                
                case_comments = ''
                if "Case Comments" in list_h and not pd.isna(row["Case Comments"]):
                    case_comments = str(str(row["Case Comments"]).strip()).replace("'", "")
                    if DB_ENCRYPTION == "YES":
                        case_comments = (fernet.encrypt(case_comments.encode())).decode("utf-8")

                case_last_step_completed = ''
                if "Case Last Step Completed" in list_h and not pd.isna(row["Case Last Step Completed"]):
                    case_last_step_completed = str(str(row["Case Last Step Completed"]).strip()).replace("'", "")
                    case_last_step_completed = case_last_step_completed.replace("'", "`")
                    if DB_ENCRYPTION == "YES":
                        case_last_step_completed = (fernet.encrypt(case_last_step_completed.encode())).decode("utf-8")

                case_last_step_completed_date = ''
                if "Case Last Step Completed Date" in list_h and row["Case Last Step Completed Date"] and not pd.isna(row["Case Last Step Completed Date"]):
                    case_last_step_completed_date = change_format(row["Case Last Step Completed Date"])

                case_next_step_to_be_completed = ''
                if "Case Next Step To Be Completed" in list_h and not pd.isna(row["Case Next Step To Be Completed"]):
                    case_next_step_to_be_completed = str(str(row["Case Next Step To Be Completed"]).strip()).replace("'", "")
                    if DB_ENCRYPTION == "YES":
                        case_next_step_to_be_completed = (fernet.encrypt(case_next_step_to_be_completed.encode())).decode("utf-8")
                
                case_next_step_to_be_completed_date = ''
                if "Case Next Step To Be Completed Date" in list_h and row["Case Next Step To Be Completed Date"] and not pd.isna(row["Case Next Step To Be Completed Date"]):
                    case_next_step_to_be_completed_date = change_format(row["Case Next Step To Be Completed Date"])
                
                case_priority_date = ''
                if "Case Priority Date" in list_h and row["Case Priority Date"] and not pd.isna(row["Case Priority Date"]):
                    case_priority_date = change_format(row["Case Priority Date"])

                case_priority_category = ''
                if "Case Priority Category" in list_h and not pd.isna(row["Case Priority Category"]):
                    case_priority_category = str(row["Case Priority Category"]).strip()
                    if DB_ENCRYPTION == "YES":
                        case_priority_category = (fernet.encrypt(case_priority_category.encode())).decode("utf-8")

                case_priority_country = ''
                if "Case Priority Country" in list_h and not pd.isna(row["Case Priority Country"]):
                    case_priority_country = str(row["Case Priority Country"]).strip()
                    if DB_ENCRYPTION == "YES":
                        case_priority_country = (fernet.encrypt(case_priority_country.encode())).decode("utf-8")

                case_approved_date = '' 
                if "Case Approved Date" in list_h and row["Case Approved Date"] and not pd.isna(row["Case Approved Date"]):
                    case_approved_date = change_format(row["Case Approved Date"])
                
                case_valid_from = ''
                if "Case Valid From" in list_h and row["Case Valid From"] and not pd.isna(row["Case Valid From"]):
                    case_valid_from = change_format(row["Case Valid From"])
                
                case_valid_to = ''
                if "Case Valid To" in list_h and row["Case Valid To"] and not pd.isna(row["Case Valid To"]):
                    case_valid_to = change_format(row["Case Valid To"])
                
                case_closed_date = ''
                if "Case Closed Date" in list_h and row["Case Closed Date"] and not pd.isna(row["Case Closed Date"]):
                    case_closed_date = change_format(row["Case Closed Date"])
                
                case_denied_date = ''
                if "Case Denied Date" in list_h and row["Case Denied Date"] and not pd.isna(row["Case Denied Date"]):
                    case_denied_date = change_format(row["Case Denied Date"])
                
                case_withdrawn_date = ''
                if "Case Withdrawn Date" in list_h and row["Case Withdrawn Date"] and not pd.isna(row["Case Withdrawn Date"]):
                    case_withdrawn_date = change_format(row["Case Withdrawn Date"])
                
                case_primary_attorney = ''
                if "Case Primary Attorney" in list_h and not pd.isna(row["Case Primary Attorney"]):
                    case_primary_attorney = str(str(row["Case Primary Attorney"]).strip()).replace("'", "")
                    if DB_ENCRYPTION == "YES":
                        case_primary_attorney = (fernet.encrypt(case_primary_attorney.encode())).decode("utf-8")
                
                case_reviewing_attorney = ''
                if "Case Reviewing Attorney" in list_h and not pd.isna(row["Case Reviewing Attorney"]):
                    case_reviewing_attorney = str(str(row["Case Reviewing Attorney"]).strip()).replace("'", "")
                    if DB_ENCRYPTION == "YES":
                        case_reviewing_attorney = (fernet.encrypt(case_reviewing_attorney.encode())).decode("utf-8")
                
                case_primary_case_manager = ''
                if "Case Primary Case Manager" in list_h and not pd.isna(row["Case Primary Case Manager"]):
                    case_primary_case_manager = str(str(row["Case Primary Case Manager"]).strip()).replace("'", "")
                    if DB_ENCRYPTION == "YES":
                        case_primary_case_manager = (fernet.encrypt(case_primary_case_manager.encode())).decode("utf-8")
                
                petition_xref = ''
                if "Petition Xref" in list_h and not pd.isna(row["Petition Xref"]):
                    petition_xref = str(row["Petition Xref"]).strip()
                
                case_id = ''
                ##print('cx ', case_xref)
                if case_xref:
                    
                    ##print("SELECT * FROM [dbo].[Case] where BeneficiaryId='{}' and CaseXref='{}' and from_name='{}'".format(beneficiary_id, case_xref, from_name))
                    results = cursor.execute("SELECT * FROM [dbo].[Case] where BeneficiaryId='{}' and CaseXref='{}' and from_name='{}'".format(beneficiary_id, case_xref, from_name)).fetchall()
                    length = len(results)
                    if length <= 0:
                        cursor.execute("INSERT INTO [dbo].[Case](CaseXref, BeneficiaryId, SourceCreatedDate, CasePetitionName, CaseType, CaseDescription, CaseFiledDate, ReceiptNumber, ReceiptStatus, RFEAuditReceivedDate,RFEAuditDueDate, RFEAuditSubmittedDate, PrimaryCaseStatus, SecondaryCaseStatus, CaseComments, LastStepCompleted, LastStepCompletedDate, NextStepAction, NextStepActionDueDate, PriorityDate, PriorityCategory, PriorityCountry, CaseApprovedDate, CaseValidFromDate, CaseExpirationDate, CaseClosedDate, CaseDeniedDate, CaseWithdrawnDate, CasePrimaryAttorney, CaseReviewingAttorney, CasePrimaryCaseManager, PetitionXref, from_name) VALUES ('{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}','{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}')".format(case_xref, beneficiary_id, case_creation_date, case_petition_name, case_type, case_description, case_filed_date, case_receipt_number, case_receipt_status, rfe_audit_received_date, rfe_audit_due_date, rfe_audit_submitted_date, primary_case_status, secondary_case_status, case_comments, case_last_step_completed, case_last_step_completed_date, case_next_step_to_be_completed, case_next_step_to_be_completed_date, case_priority_date, case_priority_category, case_priority_country, case_approved_date, case_valid_from, case_valid_to, case_closed_date, case_denied_date, case_withdrawn_date, case_primary_attorney, case_reviewing_attorney, case_primary_case_manager, petition_xref, from_name))
                        cursor.execute("SELECT @@IDENTITY AS ID;")
                        case_id = cursor.fetchone()[0]
                        cursor.commit()
                    else:
                        case_id = results[0].CaseId
                        cursor.execute("UPDATE [dbo].[Case] SET CaseXref='{}', BeneficiaryId='{}', SourceCreatedDate='{}', CasePetitionName='{}', CaseType='{}', CaseDescription='{}', CaseFiledDate='{}', ReceiptNumber='{}', ReceiptStatus='{}', RFEAuditReceivedDate='{}', RFEAuditDueDate='{}', RFEAuditSubmittedDate='{}', PrimaryCaseStatus='{}', SecondaryCaseStatus='{}', CaseComments='{}', LastStepCompleted='{}', LastStepCompletedDate='{}', NextStepAction='{}', NextStepActionDueDate='{}', PriorityDate='{}', PriorityCategory='{}', PriorityCountry='{}', CaseApprovedDate='{}', CaseValidFromDate='{}', CaseExpirationDate='{}', CaseClosedDate='{}', CaseDeniedDate='{}', CaseWithdrawnDate='{}', CasePrimaryAttorney='{}', CaseReviewingAttorney='{}', CasePrimaryCaseManager='{}', PetitionXref='{}', from_name='{}' WHERE CaseId='{}'".format(case_xref, beneficiary_id, case_creation_date, case_petition_name, case_type, case_description, case_filed_date, case_receipt_number, case_receipt_status, rfe_audit_received_date, rfe_audit_due_date, rfe_audit_submitted_date, primary_case_status, secondary_case_status, case_comments, case_last_step_completed, case_last_step_completed_date, case_next_step_to_be_completed, case_next_step_to_be_completed_date, case_priority_date, case_priority_category, case_priority_country, case_approved_date, case_valid_from, case_valid_to, case_closed_date, case_denied_date, case_withdrawn_date, case_primary_attorney, case_reviewing_attorney, case_primary_case_manager, petition_xref, from_name, case_id))
                        cursor.commit()



                
def generate_report():
    
    fernet_key = b'zJD8OVkFNpd5N4fJw6pqaWiDrvybkselSQ0fF9SwXfw='
    fernet = Fernet(fernet_key)
    current_time = datetime.now() 
    month = str(current_time.month).rjust(2, '0')
    day = str(current_time.day).rjust(2, '0')
    todate = month+''+day+''+str(current_time.year)

    case_data_confirm = False
    resultsb = cursor.execute("SELECT c.* FROM dbo.[Case] as c LEFT JOIN dbo.Beneficiary as b ON b.BeneficiaryId=c.BeneficiaryId LEFT JOIN dbo.Petitioner as p ON p.PetitionerId=b.PetitionerId ").fetchall()
    lengthb = len(resultsb)
    if lengthb > 0:
        case_data_confirm = True
    
    result_filepath = 'GT Audit Report_'+str(todate)+'.xlsx'
    result_filepath_folder = 'Processed Reports/'+str(result_filepath)
    process_report(result_filepath_folder, case_data_confirm)
    print('Audit Report Generated')
    if SEND_MAIL == "YES":
        print('Sending Mail')
        send_email(result_filepath)
    
    print('')


def process_report(result_filepath, case_data_confirm):
   
    writer = pd.ExcelWriter(result_filepath, engine='xlsxwriter', date_format='m/d/yyyy')
    

    #Tab 1

    headers2 = ['Beneficiary Id', 'Organization Name', 'Petitioner Name', 'Petitioner of Primary Beneficiary', 'Beneficiary Type (Employee / Dependent)', 'Beneficiary Record Creation Date', 'Beneficiary Record Status', 'Beneficiary Record Inactivation Date', 'Beneficiary Employee Id', 'Beneficiary Last Name', 'Beneficiary First Name', 'Beneficiary Middle Name', 'Primary Beneficiary Id', 'Primary Beneficiary Last Name', 'Primary Beneficiary First Name', 'Relation', 'Immigration Status (I-94)', 'Immigration Status Expiration Date (I-94)', 'I-797 Approved Date', 'I-797 Status', 'I-797 Expiration Date', 'Final NIV (Maxout) Date', 'Maxout Date Applicability and Note', 'Case Id', 'Case Created Date', 'Case Petition Name', 'Case Description', 'Case Filed Date', 'Case Receipt Number', 'Case Receipt Status', 'RFE/Audit Received Date', 'RFE/Audit Response Due Date', 'RFE/Audit Response Submitted Date', 'Primary Case Status', 'Secondary Case Status', 'Case Comments', 'Case Last Step Completed', 'Case Last Step Completed Date', 'Case Next Step To Be Completed', 'Case Next Step To Be Completed Date', 'Case Priority Date', 'Case Priority Category', 'Case Priority Country', 'Case Approved Date', 'Case Valid From', 'Case Valid To', 'Case Closed Date', 'Case Denied Date', 'Case Withdrawn Date', 'GT Representative']

    headers_table2 = ['BeneficiaryXref','OrganizationName','PetitionerName','Primary_Petitioner','BeneficiaryType','SourceCreatedDate','BeneficiaryRecordStatus','InactiveDate','EmployeeId','LastName','FirstName','MiddleName','PrimaryBeneficiaryXref','PrimaryBeneficiaryLastName','PrimaryBeneficiaryFirstName','RelationType','ImmigrationStatus','ImmigrationStatusExpirationDate','MostRecentI797IssueApprovalDate','MostRecentI797Status','I797ExpirationDate','FinalNivDate','MaxOutDateNote','CaseXref','CaseSourceCreatedDate','CasePetitionName', 'CaseDescription','CaseFiledDate','ReceiptNumber','ReceiptStatus','RFEAuditReceivedDate','RFEAuditDueDate','RFEAuditSubmittedDate', 'PrimaryCaseStatus','SecondaryCaseStatus','CaseComments','LastStepCompleted','LastStepCompletedDate','NextStepAction','NextStepActionDueDate','PriorityDate','PriorityCategory','PriorityCountry','CaseApprovedDate','CaseValidFromDate','CaseExpirationDate','CaseClosedDate','CaseDeniedDate','CaseWithdrawnDate','CasePrimaryCaseManager']

    date_columns2 = ['SourceCreatedDate', 'InactiveDate', 'ImmigrationStatusExpirationDate', 'MostRecentI797IssueApprovalDate', 'I797ExpirationDate', 'FinalNivDate', 'CaseSourceCreatedDate', 'CaseFiledDate','RFEAuditReceivedDate','RFEAuditDueDate','RFEAuditSubmittedDate', 'LastStepCompletedDate','NextStepActionDueDate','PriorityDate', 'CaseApprovedDate','CaseValidFromDate','CaseExpirationDate','CaseClosedDate','CaseDeniedDate','CaseWithdrawnDate']
        
    
    results_active2 = cursor.execute("SELECT b.*,c.CaseXref,c.SourceCreatedDate as CaseSourceCreatedDate,c.CasePetitionName,\
        c.CaseFiledDate, c.ReceiptNumber,c.ReceiptStatus, c.RFEAuditReceivedDate,c.RFEAuditDueDate, c.RFEAuditSubmittedDate, \
        c.PrimaryCaseStatus, c.SecondaryCaseStatus, c.CaseComments, c.LastStepCompleted, c.LastStepCompletedDate, c.NextStepAction, c.NextStepActionDueDate, c.PriorityDate, c.PriorityCategory, c.PriorityCountry, c.CaseApprovedDate, c.CaseValidFromDate, c.CaseExpirationDate, c.CaseClosedDate, c.CaseDeniedDate, c.CaseWithdrawnDate, c.CasePrimaryCaseManager,\
        o.OrganizationName,p.PetitionerName,p2.PetitionerName as Primary_Petitioner,\
        CASE WHEN b.IsActive=1 THEN 'Active' ELSE 'Inactive' END as BeneficiaryRecordStatus, \
        e.EmployeeId, c.CaseDescription \
        FROM dbo.[Case] as c  \
        LEFT JOIN  dbo.Beneficiary as b on c.BeneficiaryId=b.BeneficiaryId\
        LEFT JOIN dbo.Petitioner as p on b.PetitionerId=p.PetitionerId\
        LEFT JOIN dbo.Organization as o on p.OrganizationId=o.OrganizationId\
        LEFT JOIN dbo.Beneficiary as b2 on b.PrimaryBeneficiaryXref=b2.BeneficiaryXref\
        LEFT JOIN dbo.Petitioner as p2 on b2.PetitionerId=p2.PetitionerId\
        LEFT JOIN dbo.Organization as o2 on p2.OrganizationId=o2.OrganizationId\
        LEFT JOIN dbo.BeneficiaryEmployment as e on e.BeneficiaryId=b.BeneficiaryId\
        where b.IsActive = '1' and c.CasePetitionName !='LCA & PAF'  and c.PrimaryCaseStatus='open' \
        ORDER BY b.is_primary_beneficiary DESC, b.LastName ASC  ").fetchall()

    results_active_qry2 = "SELECT b.*,c.CaseXref,c.SourceCreatedDate as CaseSourceCreatedDate,c.CasePetitionName,\
        c.CaseFiledDate, c.ReceiptNumber,c.ReceiptStatus, c.RFEAuditReceivedDate,c.RFEAuditDueDate, c.RFEAuditSubmittedDate, \
        c.PrimaryCaseStatus, c.SecondaryCaseStatus, c.CaseComments, c.LastStepCompleted, c.LastStepCompletedDate, c.NextStepAction, c.NextStepActionDueDate, c.PriorityDate, c.PriorityCategory, c.PriorityCountry, c.CaseApprovedDate, c.CaseValidFromDate, c.CaseExpirationDate, c.CaseClosedDate, c.CaseDeniedDate, c.CaseWithdrawnDate, c.CasePrimaryCaseManager,\
        o.OrganizationName,p.PetitionerName,p2.PetitionerName as Primary_Petitioner,\
        CASE WHEN b.IsActive=1 THEN 'Active' ELSE 'Inactive' END as BeneficiaryRecordStatus, \
        e.EmployeeId, c.CaseDescription \
        FROM dbo.[Case] as c  \
        LEFT JOIN  dbo.Beneficiary as b on c.BeneficiaryId=b.BeneficiaryId\
        LEFT JOIN dbo.Petitioner as p on b.PetitionerId=p.PetitionerId\
        LEFT JOIN dbo.Organization as o on p.OrganizationId=o.OrganizationId\
        LEFT JOIN dbo.Beneficiary as b2 on b.PrimaryBeneficiaryXref=b2.BeneficiaryXref\
        LEFT JOIN dbo.Petitioner as p2 on b2.PetitionerId=p2.PetitionerId\
        LEFT JOIN dbo.Organization as o2 on p2.OrganizationId=o2.OrganizationId\
        LEFT JOIN dbo.BeneficiaryEmployment as e on e.BeneficiaryId=b.BeneficiaryId\
        where b.IsActive = '1' and c.CasePetitionName !='LCA & PAF'  and c.PrimaryCaseStatus='open' \
        ORDER BY b.is_primary_beneficiary DESC, b.LastName ASC  "
    
    
    df2 = pd.read_sql(results_active_qry2, conn)
    for dfcol in df2.columns:
        if dfcol not in headers_table2:
            #df2.pop(dfcol)
            df2.drop(dfcol, axis=1, inplace=True)

    # altering the DataFrame - Column order
    df2 = df2[headers_table2]
    for d_h in date_columns2:
        if d_h in df2:
            if "1900-01-01" in df2[d_h]:
                df2[d_h] = ""
            else:
                df2[d_h] = pd.to_datetime(df2[d_h], format='%Y-%m-%d', errors='coerce').dt.date
    df2.columns = headers2 #changing dataframe all column names

    df2.to_excel(writer, "Open Cases", startrow=0, columns=headers2, index=False)

    
    #tab 2
    headers3 = ['Beneficiary Id', 'Organization Name', 'Petitioner Name', 'Petitioner of Primary Beneficiary', 'Beneficiary Type (Employee / Dependent)', 'Beneficiary Record Creation Date', 'Beneficiary Record Status', 'Beneficiary Record Inactivation Date', 'Beneficiary Employee Id', 'Beneficiary Last Name', 'Beneficiary First Name', 'Beneficiary Middle Name', 'Primary Beneficiary Id', 'Primary Beneficiary Last Name', 'Primary Beneficiary First Name', 'Relation', 'Immigration Status (I-94)', 'Immigration Status Expiration Date (I-94)', 'I-797 Approved Date', 'I-797 Status', 'I-797 Expiration Date', 'Final NIV (Maxout) Date', 'Maxout Date Applicability and Note', 'Case Id', 'Case Created Date', 'Case Petition Name', 'Case Description', 'Case Filed Date', 'Case Receipt Number', 'Case Receipt Status', 'RFE/Audit Received Date', 'RFE/Audit Response Due Date', 'RFE/Audit Response Submitted Date', 'Primary Case Status', 'Secondary Case Status', 'Case Comments', 'Case Last Step Completed', 'Case Last Step Completed Date', 'Case Next Step To Be Completed', 'Case Next Step To Be Completed Date', 'Case Priority Date', 'Case Priority Category', 'Case Priority Country', 'Case Approved Date', 'Case Valid From', 'Case Valid To', 'Case Closed Date', 'Case Denied Date', 'Case Withdrawn Date', 'GT Representative']

    headers_table3 = ['BeneficiaryXref','OrganizationName','PetitionerName','Primary_Petitioner','BeneficiaryType','SourceCreatedDate','BeneficiaryRecordStatus','InactiveDate','EmployeeId','LastName','FirstName','MiddleName','PrimaryBeneficiaryXref','PrimaryBeneficiaryLastName','PrimaryBeneficiaryFirstName','RelationType','ImmigrationStatus','ImmigrationStatusExpirationDate','MostRecentI797IssueApprovalDate','MostRecentI797Status','I797ExpirationDate','FinalNivDate','MaxOutDateNote','CaseXref','CaseSourceCreatedDate','CasePetitionName', 'CaseDescription','CaseFiledDate','ReceiptNumber','ReceiptStatus','RFEAuditReceivedDate','RFEAuditDueDate','RFEAuditSubmittedDate', 'PrimaryCaseStatus','SecondaryCaseStatus','CaseComments','LastStepCompleted','LastStepCompletedDate','NextStepAction','NextStepActionDueDate','PriorityDate','PriorityCategory','PriorityCountry','CaseApprovedDate','CaseValidFromDate','CaseExpirationDate','CaseClosedDate','CaseDeniedDate','CaseWithdrawnDate','CasePrimaryCaseManager']

    date_columns3 = ['SourceCreatedDate','InactiveDate','ImmigrationStatusExpirationDate','MostRecentI797IssueApprovalDate','I797ExpirationDate','FinalNivDate','CaseSourceCreatedDate','CaseFiledDate','RFEAuditReceivedDate','RFEAuditDueDate','RFEAuditSubmittedDate', 'LastStepCompletedDate','NextStepActionDueDate','PriorityDate','CaseApprovedDate','CaseValidFromDate','CaseExpirationDate','CaseClosedDate','CaseDeniedDate','CaseWithdrawnDate']
    
    #chk_months = date.today() + relativedelta(months=+2) #with in 3 months including current month
    chk_months = date.today() - relativedelta(days=89)
    to_date= str(chk_months)
    from_date = str(date.today())
    
    results_active3 = cursor.execute("SELECT b.*,c.CaseXref,c.SourceCreatedDate as CaseSourceCreatedDate,c.CasePetitionName,\
        c.CaseFiledDate, c.ReceiptNumber,c.ReceiptStatus, c.RFEAuditReceivedDate,c.RFEAuditDueDate, c.RFEAuditSubmittedDate, \
        c.PrimaryCaseStatus, c.SecondaryCaseStatus, c.CaseComments, c.LastStepCompleted, c.LastStepCompletedDate, c.NextStepAction, c.NextStepActionDueDate, c.PriorityDate, c.PriorityCategory, c.PriorityCountry, c.CaseApprovedDate, c.CaseValidFromDate, c.CaseExpirationDate, c.CaseClosedDate, c.CaseDeniedDate, c.CaseWithdrawnDate, c.CasePrimaryCaseManager,\
        o.OrganizationName,p.PetitionerName,p2.PetitionerName as Primary_Petitioner,\
        CASE WHEN b.IsActive=1 THEN 'Active' ELSE 'Inactive' END as BeneficiaryRecordStatus, \
        e.EmployeeId, c.CaseDescription\
        FROM dbo.[Case] as c  \
        LEFT JOIN  dbo.Beneficiary as b on c.BeneficiaryId=b.BeneficiaryId\
        LEFT JOIN dbo.Petitioner as p on b.PetitionerId=p.PetitionerId\
        LEFT JOIN dbo.Organization as o on p.OrganizationId=o.OrganizationId\
        LEFT JOIN dbo.Beneficiary as b2 on b.PrimaryBeneficiaryXref=b2.BeneficiaryXref\
        LEFT JOIN dbo.Petitioner as p2 on b2.PetitionerId=p2.PetitionerId\
        LEFT JOIN dbo.Organization as o2 on p2.OrganizationId=o2.OrganizationId\
        LEFT JOIN dbo.BeneficiaryEmployment as e on e.BeneficiaryId=b.BeneficiaryId\
        where (b.IsActive = '1')  and c.CasePetitionName !='LCA & PAF'  and (c.PrimaryCaseStatus='approved' or PrimaryCaseStatus='Approved' or c.PrimaryCaseStatus='closed')\
        and ((c.CaseApprovedDate BETWEEN '{}' AND '{}' ) OR \
        (c.CaseClosedDate BETWEEN '{}' AND '{}' ))\
        ORDER BY b.is_primary_beneficiary DESC, b.LastName ASC ".format(to_date, from_date, to_date, from_date)).fetchall()
    
    results_active_qry3 = "SELECT b.*,c.CaseXref,c.SourceCreatedDate as CaseSourceCreatedDate,c.CasePetitionName,\
        c.CaseFiledDate, c.ReceiptNumber,c.ReceiptStatus, c.RFEAuditReceivedDate,c.RFEAuditDueDate, c.RFEAuditSubmittedDate, \
        c.PrimaryCaseStatus, c.SecondaryCaseStatus, c.CaseComments, c.LastStepCompleted, c.LastStepCompletedDate, c.NextStepAction, c.NextStepActionDueDate, c.PriorityDate, c.PriorityCategory, c.PriorityCountry, c.CaseApprovedDate, c.CaseValidFromDate, c.CaseExpirationDate, c.CaseClosedDate, c.CaseDeniedDate, c.CaseWithdrawnDate, c.CasePrimaryCaseManager,\
        o.OrganizationName,p.PetitionerName,p2.PetitionerName as Primary_Petitioner,\
        CASE WHEN b.IsActive=1 THEN 'Active' ELSE 'Inactive' END as BeneficiaryRecordStatus, \
        e.EmployeeId, c.CaseDescription\
        FROM dbo.[Case] as c  \
        LEFT JOIN  dbo.Beneficiary as b on c.BeneficiaryId=b.BeneficiaryId\
        LEFT JOIN dbo.Petitioner as p on b.PetitionerId=p.PetitionerId\
        LEFT JOIN dbo.Organization as o on p.OrganizationId=o.OrganizationId\
        LEFT JOIN dbo.Beneficiary as b2 on b.PrimaryBeneficiaryXref=b2.BeneficiaryXref\
        LEFT JOIN dbo.Petitioner as p2 on b2.PetitionerId=p2.PetitionerId\
        LEFT JOIN dbo.Organization as o2 on p2.OrganizationId=o2.OrganizationId\
        LEFT JOIN dbo.BeneficiaryEmployment as e on e.BeneficiaryId=b.BeneficiaryId\
        where (b.IsActive = '1')  and c.CasePetitionName !='LCA & PAF'  and (c.PrimaryCaseStatus='approved' or PrimaryCaseStatus='Approved' or c.PrimaryCaseStatus='closed') \
        and ((c.CaseApprovedDate BETWEEN '{}' AND '{}' ) OR \
        (c.CaseClosedDate BETWEEN '{}' AND '{}' ))\
        ORDER BY b.is_primary_beneficiary DESC, b.LastName ASC ".format(to_date, from_date, to_date, from_date)
    
    #print(results_active_qry3)
    df3 = pd.read_sql(results_active_qry3, conn)
    for dfcol in df3.columns:
        if dfcol not in headers_table3:
            df3.drop(dfcol, axis=1, inplace=True)

    # altering the DataFrame - Column order
    df3 = df3[headers_table3]
    for d_h in date_columns3:
        if d_h in df3:
            if "1900-01-01" in df3[d_h]:
                df3[d_h] = ""
            else:
                df3[d_h] = pd.to_datetime(df3[d_h], format='%Y-%m-%d', errors='coerce').dt.date
    df3.columns = headers3 #changing dataframe all column names
    df3.to_excel(writer, "Approved & Closed Cases", startrow=0, columns=headers3, index=False)

    #Tab 6
    headers = ['Beneficiary Id', 'Organization Name', 'Petitioner Name','Petitioner of Primary Beneficiary','Beneficiary Type (Employee / Dependent)', 'Beneficiary Record Creation Date', 'Beneficiary Record Status', 'Beneficiary Record Inactivation Date', 'Beneficiary Employee Id', 'Beneficiary Last Name', 'Beneficiary First Name', 'Beneficiary Middle Name', 'Primary Beneficiary Id', 'Primary Beneficiary Last Name', 'Primary Beneficiary First Name', 'Relation', 'Country of Birth', 'Country of Citizenship', 'Immigration Status (I-94)', 'Immigration Status Expiration Date (I-94)', 'I-797 Status', 'I-797 Expiration Date', 'Final NIV (Maxout) Date', 'Maxout Date Applicability and Note', 'PED', 'EAD Type', 'EAD Expiration Date', 'AP Expiration Date', 'DS-2019 Valid From', 'DS-2019 Expiration Date', 'Re-Entry Permit Expiration Date', 'Green Card Expiration Date', 'Passport Expiration Date', 'Visa Type', 'Visa Expiration Date', 'Employee Hire Date', 'Current Job Title', 'Work Address-Street', 'Work Address-City', 'Work Address-State', 'Work Address-Zip', 'Work Address-Country']

    #, 'Priority1Date', 'Priority1Category', 'Priority1Country', '', '', 'Priority2Date', 'Priority2Category', 'Priority2Country', '', '','Priority3Date', 'Priority3Category', 'Priority3Country', '', '','Priority4Date', 'Priority4Category', 'Priority4Country', '', '','Priority5Date', 'Priority5Category', 'Priority5Country', '', ''

    #, 'EADAPType', 'EadApExpirationDate'
    headers_table = ['BeneficiaryXref', 'OrganizationName', 'PetitionerName','Primary_Petitioner','BeneficiaryType', 'SourceCreatedDate', 'BeneficiaryRecordStatus', 'InactiveDate', 'EmployeeId', 'LastName', 'FirstName', 'MiddleName', 'PrimaryBeneficiaryXref', 'PrimaryBeneficiaryLastName', 'PrimaryBeneficiaryFirstName', 'RelationType', 'BirthCountry', 'CitizenshipCountry', 'ImmigrationStatus', 'ImmigrationStatusExpirationDate', 'MostRecentI797Status', 'I797ExpirationDate', 'FinalNivDate', 'MaxOutDateNote', 'VisaPedDate', 'EadType', 'EadExpirationDate', 'AdvanceParoleExpirationDate', 'Ds2019ValidFromDate', 'Ds2019ExpirationDate', 'ReEntryPermitExpirationDate', 'GreenCardExpirationDate', 'MostRecentPassportExpirationDate', 'VisaType', 'VisaExpirationDate', 'HireDate', 'JobTitle', 'Address1', 'City', 'StateProvince', 'ZipCode', 'Country']

    date_columns1 = ['SourceCreatedDate', 'InactiveDate', 'ImmigrationStatusExpirationDate', 'I797ExpirationDate', 'FinalNivDate',  'VisaPedDate',  'EadExpirationDate', 'AdvanceParoleExpirationDate', 'Ds2019ValidFromDate', 'Ds2019ExpirationDate', 'ReEntryPermitExpirationDate', 'GreenCardExpirationDate', 'MostRecentPassportExpirationDate', 'VisaExpirationDate', 'HireDate']
    
    

    results_active = cursor.execute("SELECT b.*,o.OrganizationName,p.PetitionerName,p2.PetitionerName as Primary_Petitioner,\
        CASE WHEN b.IsActive=1 THEN 'Active' ELSE 'Inactive' END as BeneficiaryRecordStatus, \
        e.EmployeeId, e.HireDate, e.JobTitle, e.Address1, e.City, e.StateProvince, e.ZipCode, e.Country,\
        bp.Priority1Date, bp.Priority1Category, bp.Priority1Country,\
        bp.Priority2Date, bp.Priority2Category, bp.Priority2Country,\
        bp.Priority3Date, bp.Priority3Category, bp.Priority3Country,\
        bp.Priority4Date, bp.Priority4Category, bp.Priority4Country,\
        bp.Priority5Date, bp.Priority5Category, bp.Priority5Country\
        FROM dbo.Beneficiary as b \
        LEFT JOIN dbo.Petitioner as p on b.PetitionerId=p.PetitionerId\
        LEFT JOIN dbo.Organization as o on p.OrganizationId=o.OrganizationId\
        LEFT JOIN dbo.Beneficiary as b2 on b.PrimaryBeneficiaryXref=b2.BeneficiaryXref\
        LEFT JOIN dbo.Petitioner as p2 on b2.PetitionerId=p2.PetitionerId\
        LEFT JOIN dbo.Organization as o2 on p2.OrganizationId=o2.OrganizationId\
        LEFT JOIN dbo.BeneficiaryEmployment as e on e.BeneficiaryId=b.BeneficiaryId\
        LEFT JOIN dbo.BeneficiaryPriorityDate as bp on bp.BeneficiaryId=b.BeneficiaryId\
        where b.IsActive = '1'  and b.from_source='beneficiary'  \
        ORDER BY b.is_primary_beneficiary DESC, b.LastName ASC").fetchall()

    results_active_qry ="SELECT b.*,o.OrganizationName,p.PetitionerName,p2.PetitionerName as Primary_Petitioner,\
        CASE WHEN b.IsActive=1 THEN 'Active' ELSE 'Inactive' END as BeneficiaryRecordStatus, \
        e.EmployeeId, e.HireDate, e.JobTitle, e.Address1, e.City, e.StateProvince, e.ZipCode, e.Country,\
        bp.Priority1Date, bp.Priority1Category, bp.Priority1Country,\
        bp.Priority2Date, bp.Priority2Category, bp.Priority2Country,\
        bp.Priority3Date, bp.Priority3Category, bp.Priority3Country,\
        bp.Priority4Date, bp.Priority4Category, bp.Priority4Country,\
        bp.Priority5Date, bp.Priority5Category, bp.Priority5Country\
        FROM dbo.Beneficiary as b \
        LEFT JOIN dbo.Petitioner as p on b.PetitionerId=p.PetitionerId\
        LEFT JOIN dbo.Organization as o on p.OrganizationId=o.OrganizationId\
        LEFT JOIN dbo.Beneficiary as b2 on b.PrimaryBeneficiaryXref=b2.BeneficiaryXref\
        LEFT JOIN dbo.Petitioner as p2 on b2.PetitionerId=p2.PetitionerId\
        LEFT JOIN dbo.Organization as o2 on p2.OrganizationId=o2.OrganizationId\
        LEFT JOIN dbo.BeneficiaryEmployment as e on e.BeneficiaryId=b.BeneficiaryId\
        LEFT JOIN dbo.BeneficiaryPriorityDate as bp on bp.BeneficiaryId=b.BeneficiaryId\
        where b.IsActive = '1'  and b.from_source='beneficiary'  \
        ORDER BY b.is_primary_beneficiary DESC, b.LastName ASC"

    df = pd.read_sql(results_active_qry, conn)
    #print(df)
    for dfcol in df.columns:
        if dfcol not in headers_table:
            df.drop(dfcol, axis=1, inplace=True)
    
    # altering the DataFrame - Column order
    df = df[headers_table]

    for d_h in date_columns1:
        if d_h in df:
            if "1900-01-01" in df[d_h]:
                df[d_h] = ""
            else:
                df[d_h] = pd.to_datetime(df[d_h], format='%Y-%m-%d', errors='coerce').dt.date
    df.columns = headers #changing dataframe all column names
    
    df.to_excel(writer, "Active Beneficiary List", startrow=0, columns=headers, index=False)
    
    writer.save()


    #Styling
    

    #Tab 1
    wb_pyxl = load_workbook(result_filepath)  
    wb_pyxl.active = 0 #active first sheet
    sheet = wb_pyxl.active
    for hdr in headers_table2:
        col = headers_table2.index(hdr)
        sheet.cell(row=1, column=col+1).alignment=Alignment(wrap_text=True, horizontal="justify", vertical="justify")
        sheet.column_dimensions[get_column_letter(col+1)].width = 14
    
    if sheet.max_row > 1:
        table = Table(displayName="Table_2", ref="A1:" + get_column_letter(sheet.max_column) + str(sheet.max_row))
    else:
        table = Table(displayName="Table_2", ref="A1:" + get_column_letter(sheet.max_column) + str(2))
        
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    sheet.add_table(table)
    
    if len(results_active2) > 0:
        for _key, s in enumerate(results_active2):
            num = _key+1
            is_primary_beneficiary = getattr(s, 'is_primary_beneficiary')
            for hdr in headers_table2:
                col = headers_table2.index(hdr)
                
                if hdr:
                    value_obj = getattr(s, hdr)
                else: 
                    value_obj = ''

                
                ##print(int(num)+1, int(col)+1, hdr, value_obj)
                
                if hdr == "SourceCreatedDate" or hdr == "InactiveDate" or hdr == "ImmigrationStatusExpirationDate" or hdr == 'MostRecentI797IssueApprovalDate' or hdr == "I797ExpirationDate" or hdr == "FinalNivDate" or hdr == "CaseSourceCreatedDate" or hdr == "CaseFiledDate" or hdr == "RFEAuditReceivedDate" or hdr == "RFEAuditDueDate" or hdr == "RFEAuditSubmittedDate"  or hdr == "LastStepCompletedDate"  or hdr == "NextStepActionDueDate"  or hdr == "PriorityDate"  or hdr == "CaseApprovedDate"  or hdr == "CaseValidFromDate"  or hdr == "CaseExpirationDate" or hdr == "CaseClosedDate"  or hdr == "CaseDeniedDate"  or hdr == "CaseWithdrawnDate":
                    ##print(value_obj)
                    if hdr == "ImmigrationStatusExpirationDate":
                        if value_obj.strip() != "D/S":
                            if 'D/S' in value_obj.strip():
                                split1 = (value_obj.strip()).split('(D/S)')
                                value_obj = change_display_format(str(split1[0]).replace('00:00:00', ''))
                                value_obj = str(value_obj)+' (D/S)'
                            else:
                                value_obj = change_display_format(str(value_obj).replace('00:00:00', ''))
                            
                    else:        
                        if str(value_obj) == '1900-01-01 00:00:00' or '1900' in str(value_obj):
                            sheet.cell(row=int(num)+1, column=col+1).value = ''
                        else:
                            value_obj = change_display_format(str(value_obj).replace('00:00:00', ''))

                    if value_obj is None or value_obj == "None":
                        sheet.cell(row=int(num)+1, column=col+1).value = ''

                    
                    sheet.cell(row=int(num)+1, column=col+1).alignment=Alignment(wrap_text=True, horizontal="justify", vertical="justify")
                    sheet.cell(row=int(num)+1, column=col+1).number_format = 'mm/dd/yyyy'
                    sheet.cell(row=int(num)+1, column=col+1).font= Font(name = 'Calibri (Body)', size= 11)
                    sheet.cell(row=int(num)+1, column=col+1).border= Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                
                    
                    pass
                else:
                    if DB_ENCRYPTION == "YES" and hdr not in encrypted_cols and value_obj:
                        #print(hdr, value_obj)
                        value_obj = fernet.decrypt(value_obj.encode()).decode("utf-8")
                        #print(hdr, value_obj)
                        #print('--------------------------------')
                    
                    if hdr == "OrganizationName":
                        if is_primary_beneficiary == 0:
                            value_obj = ''
                    elif hdr == "PetitionerName":
                        if is_primary_beneficiary == 0:
                            value_obj = 'Individual'
                    elif hdr == "BeneficiaryType":
                        if value_obj.casefold() == 'Yes'.casefold():
                            value_obj = "Employee"
                        elif value_obj.casefold() == 'No'.casefold():
                            value_obj = "Dependent"

                    if value_obj is None or value_obj == "None":
                        value_obj = ''

                    sheet.cell(row=int(num)+1, column = int(col)+1).value = str(value_obj) 
                    sheet.cell(row=int(num)+1, column=col+1).alignment=Alignment(wrap_text=True, horizontal="justify", vertical="justify")
                    sheet.cell(row=int(num)+1, column=col+1).font= Font(name = 'Calibri (Body)', size= 11)
                    sheet.cell(row=int(num)+1, column=col+1).border= Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    
                    pass
    else:
        if case_data_confirm == False:
            sheet.cell(row=2, column = 1).value = "No case data in source file"
        else:
            sheet.cell(row=2, column = 1).value = "No records matching criteria"
            
    wb_pyxl.save(result_filepath)

    #Tab 3
    wb_pyxl = load_workbook(result_filepath)  
    wb_pyxl.active = 1 
    sheet = wb_pyxl.active
    for hdr in headers_table3:
        col = headers_table3.index(hdr)
        sheet.cell(row=1, column=col+1).alignment=Alignment(wrap_text=True, horizontal="justify", vertical="justify")
        sheet.column_dimensions[get_column_letter(col+1)].width = 14
    
    if sheet.max_row > 1:
        table = Table(displayName="Table_3", ref="A1:" + get_column_letter(sheet.max_column) + str(sheet.max_row))
    else:
        table = Table(displayName="Table_3", ref="A1:" + get_column_letter(sheet.max_column) + str(2))
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    sheet.add_table(table)
    
    if len(results_active3) > 0:
        for _key, s in enumerate(results_active3):
            num = _key+1
            is_primary_beneficiary = getattr(s, 'is_primary_beneficiary')
            for hdr in headers_table3:
                col = headers_table3.index(hdr)
                
                if hdr:
                    value_obj = getattr(s, hdr)
                else: 
                    value_obj = ''
                
                ##print(int(num)+1, int(col)+1, hdr, value_obj)
                
                if hdr == "SourceCreatedDate" or hdr == "InactiveDate" or hdr == "ImmigrationStatusExpirationDate" or hdr == 'MostRecentI797IssueApprovalDate' or hdr == "I797ExpirationDate" or hdr == "FinalNivDate" or hdr == "CaseSourceCreatedDate" or hdr == "CaseFiledDate" or hdr == "RFEAuditReceivedDate" or hdr == "RFEAuditDueDate" or hdr == "RFEAuditSubmittedDate"  or hdr == "LastStepCompletedDate"  or hdr == "NextStepActionDueDate"  or hdr == "PriorityDate"  or hdr == "CaseApprovedDate"  or hdr == "CaseValidFromDate"  or hdr == "CaseExpirationDate" or hdr == "CaseClosedDate"  or hdr == "CaseDeniedDate"  or hdr == "CaseWithdrawnDate":
                    ##print(value_obj)
                    if hdr == "ImmigrationStatusExpirationDate":
                        if value_obj.strip() != "D/S":
                            if 'D/S' in value_obj.strip():
                                split1 = (value_obj.strip()).split('(D/S)')
                                value_obj = change_display_format(str(split1[0]).replace('00:00:00', ''))
                                value_obj = str(value_obj)+' (D/S)'
                            else:
                                value_obj = change_display_format(str(value_obj).replace('00:00:00', ''))
                            
                    else:        
                        if str(value_obj) == '1900-01-01 00:00:00':
                            value_obj = ''
                            sheet.cell(row=int(num)+1, column=col+1).value = ''
                        else:
                            value_obj = change_display_format(str(value_obj).replace('00:00:00', ''))

                    if value_obj is None or value_obj == "None":
                        value_obj = ''
                        sheet.cell(row=int(num)+1, column=col+1).value = ''

                    
                    sheet.cell(row=int(num)+1, column=col+1).alignment=Alignment(wrap_text=True, horizontal="justify", vertical="justify")
                    sheet.cell(row=int(num)+1, column=col+1).number_format = 'mm/dd/yyyy'
                    sheet.cell(row=int(num)+1, column=col+1).font= Font(name = 'Calibri (Body)', size= 11)
                    sheet.cell(row=int(num)+1, column=col+1).border= Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                
                    
                    pass
                else:
                    if DB_ENCRYPTION == "YES" and hdr not in encrypted_cols and value_obj:
                        #print(hdr, value_obj)
                        value_obj = fernet.decrypt(value_obj.encode()).decode("utf-8")
                        #print(hdr, value_obj)
                        #print('--------------------------------')
                    
                    if hdr == "OrganizationName":
                        if is_primary_beneficiary == 0:
                            value_obj = ''
                    elif hdr == "PetitionerName":
                        if is_primary_beneficiary == 0:
                            value_obj = 'Individual'
                    elif hdr == "BeneficiaryType":
                        if value_obj.casefold() == 'Yes'.casefold():
                            value_obj = "Employee"
                        elif value_obj.casefold() == 'No'.casefold():
                            value_obj = "Dependent"

                    if value_obj is None or value_obj == "None":
                        value_obj = ''

                    sheet.cell(row=int(num)+1, column = int(col)+1).value = str(value_obj) 
                    sheet.cell(row=int(num)+1, column=col+1).alignment=Alignment(wrap_text=True, horizontal="justify", vertical="justify")
                    sheet.cell(row=int(num)+1, column=col+1).font= Font(name = 'Calibri (Body)', size= 11)
                    sheet.cell(row=int(num)+1, column=col+1).border= Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    
                    pass
    else:
        if case_data_confirm == False:
            sheet.cell(row=2, column = 1).value = "No case data in source file"
        else:
            sheet.cell(row=2, column = 1).value = "No records matching criteria"
    wb_pyxl.save(result_filepath)

    #Tab 6
    wb_pyxl = load_workbook(result_filepath)  
    wb_pyxl.active = 2 
    sheet = wb_pyxl.active
    for hdr in headers_table:
        col = headers_table.index(hdr)
        #print(hdr, col)
        sheet.cell(row=1, column=col+1).alignment=Alignment(wrap_text=True, horizontal="justify", vertical="justify")
        sheet.column_dimensions[get_column_letter(col+1)].width = 14
    
    if sheet.max_row > 1:
        table = Table(displayName="Table_1", ref="A1:" + get_column_letter(sheet.max_column) + str(sheet.max_row))
    else:
        table = Table(displayName="Table_1", ref="A1:" + get_column_letter(sheet.max_column) + str(2))
        
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    sheet.add_table(table)
    if len(results_active) > 0:
        for _key, s in enumerate(results_active):
            num = _key+1
            is_primary_beneficiary = getattr(s, 'is_primary_beneficiary')
            for hdr in headers_table:
                col = headers_table.index(hdr)
                
                if hdr:
                    value_obj = getattr(s, hdr)
                else: 
                    value_obj = ''
                
                ##print(int(num)+1, int(col)+1, hdr, value_obj)
                
                if hdr == "SourceCreatedDate" or hdr == "InactiveDate" or hdr == "ImmigrationStatusExpirationDate" or hdr == "I797ExpirationDate" or hdr == "FinalNivDate" or hdr == "VisaPedDate" or hdr == "EadExpirationDate" or hdr == "AdvanceParoleExpirationDate" or hdr == "EadApExpirationDate" or hdr == "Ds2019ValidFromDate"  or hdr == "Ds2019ExpirationDate"  or hdr == "ReEntryPermitExpirationDate"  or hdr == "GreenCardExpirationDate"  or hdr == "MostRecentPassportExpirationDate"  or hdr == "VisaExpirationDate"  or hdr == "HireDate" or hdr == "Priority1Date"  or hdr == "Priority2Date"  or hdr == "Priority3Date"  or hdr == "Priority4Date"  or hdr == "Priority5Date"  :
                    ##print(value_obj)
                    if hdr == "ImmigrationStatusExpirationDate":
                        if value_obj.strip() != "D/S":
                            if 'D/S' in value_obj.strip():
                                split1 = (value_obj.strip()).split('(D/S)')
                                value_obj = change_display_format(str(split1[0]).replace('00:00:00', ''))
                                value_obj = str(value_obj)+' (D/S)'
                            else:
                                value_obj = change_display_format(str(value_obj).replace('00:00:00', ''))
                            
                    else:        
                        if str(value_obj) == '1900-01-01 00:00:00' or '1900' in str(value_obj):
                            sheet.cell(row=int(num)+1, column=col+1).value = ''
                        else:
                            value_obj = change_display_format(str(value_obj).replace('00:00:00', ''))

                    if value_obj is None or value_obj == "None":
                        value_obj = ''

                    
                    sheet.cell(row=int(num)+1, column=col+1).alignment=Alignment(wrap_text=True, horizontal="justify", vertical="justify")
                    sheet.cell(row=int(num)+1, column=col+1).number_format = 'mm/dd/yyyy'
                    sheet.cell(row=int(num)+1, column=col+1).font= Font(name = 'Calibri (Body)', size= 11)
                    sheet.cell(row=int(num)+1, column=col+1).border= Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                
                    
                    pass
                else:

                    if DB_ENCRYPTION == "YES" and hdr not in encrypted_cols and value_obj:
                        #print(hdr, value_obj)
                        value_obj = fernet.decrypt(value_obj.encode()).decode("utf-8")
                        #print(hdr, value_obj)
                        #print('--------------------------------')
                        

                    if hdr == "OrganizationName":
                        if is_primary_beneficiary == 0:
                            value_obj = ''
                    elif hdr == "PetitionerName":
                        if is_primary_beneficiary == 0:
                            value_obj = 'Individual'
                    elif hdr == "BeneficiaryType":
                        if value_obj.casefold() == 'Yes'.casefold():
                            value_obj = "Employee"
                        elif value_obj.casefold() == 'No'.casefold():
                            value_obj = "Dependent"


                    if value_obj is None or value_obj == "None":
                        value_obj = ''

                    sheet.cell(row=int(num)+1, column = int(col)+1).value = str(value_obj) 
                    sheet.cell(row=int(num)+1, column=col+1).alignment=Alignment(wrap_text=True, horizontal="justify", vertical="justify")
                    sheet.cell(row=int(num)+1, column=col+1).font= Font(name = 'Calibri (Body)', size= 11)
                    sheet.cell(row=int(num)+1, column=col+1).border= Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    
                    pass
                
            #return False
    else:
        sheet.cell(row=2, column = 1).value = "No beneficiary data in source file"
    
        
    wb_pyxl.save(result_filepath)
    
    
    
    wb_pyxl.active = 0 #active First sheet
    wb_pyxl.save(result_filepath)



def send_email(filename):

    
    try:
        info = "Audit Report"
        subject = str(info)
        receivers_mail = ''
        receivers_cc_mail = ''
        #print(results)
        receivers_mail = ('shiv@immilytics.com').split(';')
        receivers_cc_mail = ('shiv@immilytics.com').split(';')
        
        body = """
        <html>
        <head></head>
        <body>
            <p>Please find attached the audit report.<br><br>
            Kind regards,<br>
            Reports Automation Team</p>
        </body>
        </html>
        """
        sender_email = "processautomation@immilytics.com"
        receiver_email =  ", ".join(receivers_mail)
        receiver_cc_email =  ", ".join(receivers_cc_mail)
        password = "iRPA@2020!"

        # Create a multipart message and set headers
        message = MIMEMultipart()
        message["From"] = sender_email
        message["To"] = receiver_email
        message["Subject"] = subject
        message["Cc"] = receiver_cc_email  # Recommended for mass emails

        # Add body to email
        message.attach(MIMEText(body, "html")) #html/plain

        file_list = [filename]
        

        for filename_ in file_list:    
            # Open PDF file in binary mode
            with open(os.path.join(cwd, 'Processed Reports', filename_), "rb") as attachment:
                # Add file as application/octet-stream
                # Email client can usually download this automatically as attachment
                part = MIMEBase("application", "octet-stream")
                part.set_payload(attachment.read())

            # Encode file in ASCII characters to send by email    
            encoders.encode_base64(part)

            # Add header as key/value pair to attachment part
            part.add_header(
                "Content-Disposition",
                f"attachment; filename= {filename_}",
            )

            # Add attachment to message and convert message to string
            message.attach(part)
            text = message.as_string()

        # Log in to server using secure context and send email
        context = ssl.create_default_context()
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
            server.login(sender_email, password)
            server.sendmail(sender_email, receiver_email.split(',')+receiver_cc_email.split(','), text)
            #update_log(str(client_short_name), 'Sent', filename, filename2, '' )
            print(str(info)+' Mail sent ')
    except Exception as e:
        #update_log(str(client_short_name), 'Not Sent', filename, filename2, str(e) )
        print('Error: Mail not sent ')
        print(e)
   
                
                

if __name__ == '__main__':
    start()
    generate_report()
    print('Finished')
    pass
    
    
    

